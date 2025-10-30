# Gestion_Alimentos.py - VERSIÓN CORREGIDA Y MEJORADA
import sys
import os
import bcrypt
import traceback
import logging
import re
import secrets
import string
from datetime import datetime, timedelta
from functools import wraps
import hashlib
from typing import Dict, List, Optional, Tuple, Any, Union
from decimal import Decimal, InvalidOperation

# Módulos de PyQt5
from PyQt5.QtWidgets import ( # pyright: ignore[reportMissingImports]
    QApplication, QMainWindow, QWidget, QVBoxLayout, QHBoxLayout,
    QLabel, QLineEdit, QPushButton, QTableWidget, QTableWidgetItem,
    QComboBox, QHeaderView, QMessageBox, QTabWidget, QDialog, QFormLayout,
    QTextEdit, QGroupBox, QCheckBox, QProgressBar, QSystemTrayIcon, QMenu, QAction,
    QSplashScreen, QDesktopWidget, QStyle, QToolBar, QStatusBar, QInputDialog,
    QFrame, QStackedWidget, QSizePolicy, QSpacerItem, QScrollArea, QGridLayout,
    QDialogButtonBox, QProgressDialog, QDateEdit, QSpinBox, QDoubleSpinBox
)
from PyQt5.QtCore import Qt, QTimer, QSettings, QSize, QPropertyAnimation, QEasingCurve, QParallelAnimationGroup, pyqtProperty, QDate # pyright: ignore[reportMissingImports]
from PyQt5.QtGui import ( # pyright: ignore[reportMissingImports]
    QDoubleValidator, QIntValidator, QFont, QIcon, QPixmap, QPalette, QColor,
    QFontDatabase, QLinearGradient, QPainter, QBrush, QPen, QKeySequence
)
from PyQt5.QtPrintSupport import QPrinter, QPrintDialog # pyright: ignore[reportMissingImports]
from PyQt5.QtGui import QTextDocument

# Módulos de Base de Datos y Visualización
import mysql.connector
import matplotlib
matplotlib.use('Qt5Agg')
import matplotlib.pyplot as plt
from matplotlib.backends.backend_qt5agg import FigureCanvasQTAgg as FigureCanvas

# Módulos para PDF y códigos de barras
from reportlab.lib.pagesizes import A4
from reportlab.pdfgen import canvas
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib import colors
# import barcode
# from barcode.writer import ImageWriter
from PIL import Image

# Módulo para exportación a Excel
try:
    from openpyxl import Workbook  # pyright: ignore[reportMissingImports]
    from openpyxl.styles import Font, PatternFill, Alignment  # pyright: ignore[reportMissingImports]
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False
    print("Advertencia: openpyxl no está instalado. La exportación a Excel no estará disponible.")

# ==============================================================================
# CONFIGURACIÓN EMPRESARIAL - MEJORADA
# ==============================================================================

# Configuración del logging empresarial
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s',
    filename='app_errors.log',
    filemode='a'
)

# Configuración de la aplicación
APP_NAME = "Sistema de Gestión de Alimentos Pro"
APP_VERSION = "4.0.0"
COMPANY_NAME = "Food Management Solutions"

# Paleta de colores moderna y atractiva
COLORS = {
    'primary': '#2563EB',      # Azul moderno vibrante
    'secondary': '#64748B',    # Gris azulado elegante
    'accent': '#3B82F6',       # Azul claro moderno
    'success': '#10B981',      # Verde esmeralda
    'warning': '#F59E0B',      # Amarillo ámbar
    'danger': '#EF4444',       # Rojo coral
    'light': '#F8FAFC',        # Gris muy claro moderno
    'dark': '#1E293B',         # Azul grisáceo oscuro
    'background': '#FFFFFF',   # Fondo blanco puro
    'surface': '#F1F5F9',      # Superficie gris clara moderna
    'text_primary': '#1E293B', # Texto principal azul grisáceo
    'text_secondary': '#64748B', # Texto secundario
    'border': '#E2E8F0',       # Bordes suaves
    'hover': '#F1F5F9',        # Hover suave
    'shadow': 'rgba(0, 0, 0, 0.08)' # Sombra sutil moderna
}

# Usuarios completos del sistema
USUARIOS_SISTEMA = [
    # Administradores
    ('paul', 'admin123', 'admin'),
    ('ana', 'admin123', 'admin'),
    
    # Empleados con roles específicos
    ('jhon', 'admin123', 'inventario'),
    ('yessenia', 'admin123', 'compras'),
    ('piero', 'admin123', 'finanzas'),
    ('cassandra', 'admin123', 'reportes'),
    
    # Atención al cliente
    ('vanina', 'admin123', 'atencion_cliente'),
    ('miryam', 'admin123', 'atencion_cliente'), 
    ('natalia', 'admin123', 'atencion_cliente')
]

# ==============================================================================
# EXCEPCIONES PERSONALIZADAS
# ==============================================================================

class DatabaseError(Exception):
    """Excepción para errores de base de datos."""
    pass

class SecurityError(Exception):
    """Excepción para errores de seguridad."""
    pass

class ValidationError(Exception):
    """Excepción para errores de validación."""
    pass

class BusinessRuleError(Exception):
    """Excepción para violaciones de reglas de negocio."""
    pass

# ==============================================================================
# DECORADORES UTILITARIOS MEJORADOS
# ==============================================================================

def handle_db_errors(func):
    """Decorador para manejar errores de base de datos y validación."""
    @wraps(func)
    def wrapper(*args, **kwargs):
        try:
            return func(*args, **kwargs)
        except ValidationError as e:
            logging.warning(f"Validación fallida en {func.__name__}: {e}")
            raise
        except DatabaseError as e:
            logging.error(f"Error de BD en {func.__name__}: {e}")
            raise
        except SecurityError as e:
            logging.error(f"Error de seguridad en {func.__name__}: {e}")
            raise
        except BusinessRuleError as e:
            logging.warning(f"Regla de negocio violada en {func.__name__}: {e}")
            raise
        except Exception as e:
            logging.critical(f"Error inesperado en {func.__name__}: {traceback.format_exc()}")
            raise DatabaseError(f"Error inesperado: {str(e)}")
    return wrapper

def require_permission(required_roles):
    """Decorador para verificar permisos de usuario."""
    if isinstance(required_roles, str):
        required_roles = [required_roles]
        
    def decorator(func):
        @wraps(func)
        def wrapper(self, *args, **kwargs):
            if not hasattr(self, 'user_role') or self.user_role not in required_roles:
                raise SecurityError(f"Se requieren permisos de {required_roles} para esta acción.")
            return func(self, *args, **kwargs)
        return wrapper
    return decorator

# ==============================================================================
# GESTOR DE CONFIGURACIÓN EMPRESARIAL - MEJORADO
# ==============================================================================

class ConfigManager:
    """Gestiona la configuración de la aplicación."""
    
    def __init__(self):
        self.settings = QSettings(COMPANY_NAME, APP_NAME)
        
    def get_database_config(self):
        """Obtiene la configuración de la base de datos."""
        return {
            'db_name': self.settings.value('database/name', 'GestionAlimentos'),
            'db_user': self.settings.value('database/user', 'root'),
            'db_password': self.settings.value('database/password', ''),  # Contraseña por defecto para root local
            'db_host': self.settings.value('database/host', 'localhost'),
            'db_port': self.settings.value('database/port', '3306')
        }
    
    def save_database_config(self, config):
        """Guarda la configuración de la base de datos."""
        for key, value in config.items():
            self.settings.setValue(f'database/{key}', value)
    
    def get_ui_config(self):
        """Obtiene la configuración de la interfaz."""
        return {
            'theme': self.settings.value('ui/theme', 'light'),
            'language': self.settings.value('ui/language', 'es'),
            'auto_save': self.settings.value('ui/auto_save', True, type=bool)
        }

# ==============================================================================
# GESTOR DE BASE DE DATOS EMPRESARIAL - OPTIMIZADO
# ==============================================================================

class DatabaseManager:
    """Gestiona la conexión y las operaciones de bajo nivel con MySQL."""

    def __init__(self, config_manager):
        self.config_manager = config_manager
        self.conn = None
        self._is_connected = False
        self._connection_attempts = 0
        self.max_attempts = 3

    def get_connection_config(self):
        """Obtiene la configuración de conexión."""
        return self.config_manager.get_database_config()

    def connect(self):
        """Establece y mantiene una única conexión a la base de datos."""
        if self._is_connected and self.conn and self.conn.is_connected():
            return self.conn

        try:
            config = self.get_connection_config()
            self.conn = mysql.connector.connect(
                host=config['db_host'],
                user=config['db_user'],
                password=config['db_password'],
                database=config['db_name'],
                port=int(config['db_port']),
                connection_timeout=10,
                autocommit=False,  # Control manual de transacciones
                auth_plugin='mysql_native_password'  # Plugin de autenticación explícito
            )
            self._is_connected = True
            self._connection_attempts = 0
            logging.info("Conexión a la base de datos establecida correctamente.")
            return self.conn
        except mysql.connector.Error as e:
            self._connection_attempts += 1
            error_message = f"Error al conectar con la base de datos (intento {self._connection_attempts}): {e}"
            logging.error(error_message)

            if self._connection_attempts >= self.max_attempts:
                self._is_connected = False
                raise DatabaseError(f"No se pudo conectar después de {self.max_attempts} intentos: {e}")

            self._is_connected = False
            raise DatabaseError(error_message)

    def close(self):
        """Cierra la conexión si está abierta."""
        if self.conn and self.conn.is_connected():
            self.conn.close()
            self._is_connected = False
            logging.info("Conexión a la base de datos cerrada.")

    def execute_query(self, query, params=None, fetch_one=False, fetch_all=False):
        """Ejecuta una consulta SQL reutilizando la conexión existente."""
        conn = None
        try:
            conn = self.connect()
            with conn.cursor(dictionary=True) as cur:
                cur.execute(query, params or ())
                
                if fetch_one:
                    result = cur.fetchone()
                    conn.commit()
                    return result
                if fetch_all:
                    result = cur.fetchall()
                    conn.commit()
                    return result
                
                conn.commit()
                return True
                
        except mysql.connector.Error as e:
            error_message = f"Error al ejecutar consulta SQL:\n{query}\nParámetros: {params}\nError: {e}"
            logging.error(error_message)

            if conn:
                conn.rollback()

            # Manejo de reconexión si la conexión se pierde
            if "connection" in str(e).lower() or "closed" in str(e).lower() or "2006" in str(e) or "2013" in str(e):
                self._is_connected = False
                logging.info("Intentando reconexión...")
                try:
                    conn = self.connect() # Reintenta la conexión
                    with conn.cursor(dictionary=True) as cur:
                        cur.execute(query, params or ())

                        if fetch_one:
                            result = cur.fetchone()
                            conn.commit()
                            return result
                        if fetch_all:
                            result = cur.fetchall()
                            conn.commit()
                            return result
                        conn.commit()
                        return True
                except mysql.connector.Error as retry_error:
                    raise DatabaseError(f"Error persistente en la BD: {retry_error}")

            raise DatabaseError(f"Error en la BD: {e}")
        except Exception as e:
            if conn:
                conn.rollback()
            logging.error(f"Error inesperado en execute_query: {traceback.format_exc()}")
            raise DatabaseError(f"Error inesperado: {e}")

    def test_connection(self):
        """Prueba la conexión a la base de datos."""
        try:
            conn = self.connect()
            if conn and conn.is_connected():
                return True
            return False
        except:
            return False

# ==============================================================================
# GESTOR DE SEGURIDAD EMPRESARIAL
# ==============================================================================

class SecurityManager:
    """Gestiona la seguridad de la aplicación."""
    
    def __init__(self):
        self.failed_attempts = {}
        self.max_attempts = 5
        self.lockout_time = timedelta(minutes=30)
        
    def hash_password(self, password):
        """Hashea la contraseña usando bcrypt con salt único."""
        try:
            if isinstance(password, str):
                password = password.encode('utf-8')
            salt = bcrypt.gensalt(rounds=12)
            hashed = bcrypt.hashpw(password, salt)
            return hashed.decode('utf-8')
        except Exception as e:
            logging.error(f"Error al hashear contraseña: {e}")
            raise SecurityError("Error al procesar la contraseña.")

    def check_password(self, password, hashed_password):
        """Verifica la contraseña de manera segura."""
        try:
            if isinstance(password, str):
                password = password.encode('utf-8')
            if isinstance(hashed_password, str):
                hashed_password = hashed_password.encode('utf-8')
            return bcrypt.checkpw(password, hashed_password)
        except Exception as e:
            logging.error(f"Error al verificar contraseña: {e}")
            return False

    def is_account_locked(self, username):
        """Verifica si una cuenta está bloqueada."""
        if username in self.failed_attempts:
            attempts, last_attempt = self.failed_attempts[username]
            if attempts >= self.max_attempts:
                if datetime.now() - last_attempt < self.lockout_time:
                    return True
                else:
                    del self.failed_attempts[username]
        return False

    def record_failed_attempt(self, username):
        """Registra un intento fallido de login."""
        if username not in self.failed_attempts:
            self.failed_attempts[username] = [1, datetime.now()]
        else:
            self.failed_attempts[username][0] += 1
            self.failed_attempts[username][1] = datetime.now()

    def reset_failed_attempts(self, username):
        """Resetea los intentos fallidos para un usuario."""
        if username in self.failed_attempts:
            del self.failed_attempts[username]

    def generate_secure_password(self, length=12):
        """Genera una contraseña segura."""
        characters = string.ascii_letters + string.digits + "!@#$%^&*"
        return ''.join(secrets.choice(characters) for _ in range(length))

# ==============================================================================
# LÓGICA DE NEGOCIO EMPRESARIAL COMPLETA - OPTIMIZADA
# ==============================================================================

class BusinessLogic:
    """Contiene las reglas de negocio, validaciones y lógica de la aplicación."""

    def __init__(self, db_manager, security_manager=None):
        self.db = db_manager
        self.security = security_manager or SecurityManager()
        self._proveedores_cache = None
        self._usuarios_cache = None
        self._productos_cache = None
        self._comprobantes_cache = None
        self._cache_timestamp = {}

    # ==================== VALIDACIONES EMPRESARIALES ====================
    
    def validate_email(self, email):
        """Valida el formato de email empresarial."""
        if email and not re.match(r'^[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}$', email):
            raise ValidationError("Formato de email inválido.")
        return True

    def validate_phone(self, phone):
        """Valida el formato de teléfono."""
        if phone and not re.match(r'^[\d\s\-\+\(\)]{8,20}$', phone):
            raise ValidationError("Formato de teléfono inválido.")
        return True

    def validate_ruc(self, ruc):
        """Valida formato de RUC."""
        if ruc and not re.match(r'^\d{11}$', ruc):
            raise ValidationError("RUC debe tener 11 dígitos.")
        return True

    def validate_password_strength(self, password):
        """Valida la fortaleza de la contraseña."""
        if len(password) < 8:
            raise ValidationError("La contraseña debe tener al menos 8 caracteres.")
        if not re.search(r'[A-Z]', password):
            raise ValidationError("La contraseña debe contener al menos una mayúscula.")
        if not re.search(r'[a-z]', password):
            raise ValidationError("La contraseña debe contener al menos una minúscula.")
        if not re.search(r'\d', password):
            raise ValidationError("La contraseña debe contener al menos un número.")
        return True

    def safe_decimal(self, value):
        """Convierte seguro a Decimal."""
        if value is None or value == '':
            return Decimal('0.00')
        try:
            return Decimal(str(value).replace(',', '.'))
        except (InvalidOperation, TypeError):
            raise ValidationError("Valor decimal inválido.")

    def safe_int(self, value):
        """Convierte seguro a entero."""
        if value is None or value == '':
            return 0
        try:
            return int(value)
        except (ValueError, TypeError):
            raise ValidationError("Valor entero inválido.")

    # ==================== MÉTODOS DE AUTENTICACIÓN ====================
    
    @handle_db_errors
    def login_user(self, username, password):
        """Verifica las credenciales del usuario con seguridad empresarial."""
        if self.security.is_account_locked(username):
            logging.warning(f"Intento de login en cuenta bloqueada: {username}")
            raise SecurityError("Cuenta temporalmente bloqueada por múltiples intentos fallidos.")

        if not username or not password:
            self.security.record_failed_attempt(username)
            return None, None, None, "Usuario o contraseña no proporcionados."

        user = self.db.execute_query(
            "SELECT idusuario, nombreusuario, contrasena, rol, activo FROM usuarios WHERE nombreusuario = %s",
            (username.strip().lower(),),
            fetch_one=True
        )

        if user and user['activo']:
            if self.security.check_password(password, user['contrasena']):
                self.security.reset_failed_attempts(username)
                logging.info(f"Login exitoso para: {username} ({user['rol']})")

                self.db.execute_query(
                    "UPDATE usuarios SET ultimo_login = NOW() WHERE idusuario = %s",
                    (user['idusuario'],)
                )

                self.db.execute_query(
                    "INSERT INTO audit_log (usuario, accion, detalles) VALUES (%s, %s, %s)",
                    (username, 'LOGIN_EXITOSO', f'Login desde Sistema')
                )

                return user['idusuario'], user['nombreusuario'], user['rol'], None
            else:
                self.security.record_failed_attempt(username)
                logging.warning(f"Contraseña incorrecta para: {username}")
                return None, None, None, "Contraseña incorrecta."
        else:
            self.security.record_failed_attempt(username)
            if user and not user['activo']:
                logging.warning(f"Usuario inactivo: {username}")
                return None, None, None, "Usuario inactivo."
            else:
                logging.warning(f"Usuario no encontrado: {username}")
                return None, None, None, "Usuario no encontrado."

    # ==================== MÉTODOS DE USUARIO ====================
    
    def get_all_usuarios(self, force_refresh=False):
        """Retorna todos los usuarios con cache inteligente."""
        cache_key = 'usuarios'
        if not force_refresh and self._is_cache_valid(cache_key):
            return self._usuarios_cache
            
        try:
            query = """
            SELECT idusuario, nombreusuario, rol, activo, fecha_creacion, 
                   ultimo_login, email, telefono
            FROM usuarios 
            ORDER BY nombreusuario
            """
            self._usuarios_cache = self.db.execute_query(query, fetch_all=True) or []
            self._update_cache_timestamp(cache_key)
            return self._usuarios_cache
        except Exception as e:
            logging.error(f"Error al obtener usuarios: {e}")
            return []

    # ==================== MÉTODOS DE PROVEEDORES ====================
    
    def get_all_proveedores(self, force_refresh=False):
        """Retorna todos los proveedores con cache."""
        cache_key = 'proveedores'
        if not force_refresh and self._is_cache_valid(cache_key):
            return self._proveedores_cache
            
        query = """
        SELECT idproveedor, nombre, contacto, telefono, email, 
               direccion, ruc, fecha_registro, activo
        FROM proveedores 
        ORDER BY nombre
        """
        self._proveedores_cache = self.db.execute_query(query, fetch_all=True) or []
        self._update_cache_timestamp(cache_key)
        return self._proveedores_cache

    def get_proveedor_by_id(self, idproveedor):
        """Obtiene un proveedor específico por ID."""
        query = """
        SELECT idproveedor, nombre, contacto, telefono, email, 
               direccion, ruc, fecha_registro, activo
        FROM proveedores 
        WHERE idproveedor = %s
        """
        return self.db.execute_query(query, (idproveedor,), fetch_one=True)

    @handle_db_errors
    def save_proveedor(self, proveedor_data):
        """Guarda o actualiza un proveedor."""
        if proveedor_data.get('idproveedor'):
            # Actualizar
            query = """
            UPDATE proveedores SET 
                nombre = %s, contacto = %s, telefono = %s, email = %s,
                direccion = %s, ruc = %s, activo = %s
            WHERE idproveedor = %s
            """
            params = (
                proveedor_data['nombre'], proveedor_data['contacto'],
                proveedor_data['telefono'], proveedor_data['email'],
                proveedor_data['direccion'], proveedor_data['ruc'],
                proveedor_data['activo'], proveedor_data['idproveedor']
            )
        else:
            # Insertar
            query = """
            INSERT INTO proveedores 
            (nombre, contacto, telefono, email, direccion, ruc, activo)
            VALUES (%s, %s, %s, %s, %s, %s, %s)
            """
            params = (
                proveedor_data['nombre'], proveedor_data['contacto'],
                proveedor_data['telefono'], proveedor_data['email'],
                proveedor_data['direccion'], proveedor_data['ruc'],
                proveedor_data['activo']
            )
        
        self.db.execute_query(query, params)
        self._invalidate_cache('proveedores')
        return True

    @handle_db_errors
    def delete_proveedor(self, idproveedor):
        """Elimina lógicamente un proveedor."""
        # Verificar si hay productos asociados
        productos_count = self.db.execute_query(
            "SELECT COUNT(*) as count FROM productos WHERE idproveedor = %s AND activo = TRUE",
            (idproveedor,),
            fetch_one=True
        )
        
        if productos_count and productos_count['count'] > 0:
            raise BusinessRuleError("No se puede eliminar el proveedor porque tiene productos activos asociados.")
        
        query = "UPDATE proveedores SET activo = FALSE WHERE idproveedor = %s"
        self.db.execute_query(query, (idproveedor,))
        self._invalidate_cache('proveedores')
        return True

    # ==================== MÉTODOS DE PRODUCTOS ====================

    def get_all_productos(self, force_refresh=False):
        """Retorna todos los productos con el nombre del proveedor."""
        cache_key = 'productos'
        if not force_refresh and self._is_cache_valid(cache_key):
            return self._productos_cache
            
        query = """
        SELECT 
            p.idproducto, 
            p.nombre, 
            pr.nombre as nombreproveedor, 
            p.descripcion,
            p.categoria,
            p.stock,
            p.stockminimo,
            p.precio,
            p.preciobase,
            p.fecharegistro,
            p.codigo_barras,
            p.activo
        FROM productos p
        JOIN proveedores pr ON p.idproveedor = pr.idproveedor
        WHERE p.activo = true
        ORDER BY p.nombre
        """
        self._productos_cache = self.db.execute_query(query, fetch_all=True) or []
        self._update_cache_timestamp(cache_key)
        return self._productos_cache

    # ==================== MÉTODOS DE IPC (OPTIMIZADO POR LOTES) ====================
    
    @handle_db_errors
    def aplicar_ipc(self, porcentaje, usuario):
        """Aplica variación de precios por IPC con optimización por lote."""
        try:
            porcentaje_decimal = self.safe_decimal(porcentaje)
            multiplicador = Decimal('1') + porcentaje_decimal / Decimal('100')
            
            # 1. Obtener todos los productos activos y sus precios actuales
            productos = self.db.execute_query(
                "SELECT idproducto, precio FROM productos WHERE activo = TRUE",
                fetch_all=True
            )
            
            if not productos:
                logging.info("IPC aplicado pero no hay productos activos para actualizar.")
                return True # Éxito, no hay nada que actualizar.
                
            historial_data = []
            
            # Preparar los datos del historial en memoria
            for producto in productos:
                precio_anterior = producto['precio']
                # Calcular el nuevo precio para el registro de historial
                precio_nuevo = precio_anterior * multiplicador
                
                # (idproducto, precioanterior, preccionuevo, motivo, tipocambio, usuariocambio)
                historial_data.append((
                    producto['idproducto'], 
                    precio_anterior, 
                    precio_nuevo, 
                    f'Ajuste IPC {porcentaje}%', 
                    'IPC', 
                    usuario
                ))
            
            # 2. Actualización masiva de precios (Optimización IPC)
            self.db.execute_query(
                """
                UPDATE productos 
                SET precio = precio * %s 
                WHERE activo = TRUE
                """,
                (multiplicador,)
            )

            # 3. Inserción masiva en historial (Optimización IPC)
            if historial_data:
                conn = self.db.connect()
                with conn.cursor() as cur:
                    insert_query = """
                    INSERT INTO historialprecios 
                    (idproducto, precioanterior, preccionuevo, motivo, tipocambio, usuariocambio) 
                    VALUES (%s, %s, %s, %s, %s, %s)
                    """
                    cur.executemany(insert_query, historial_data)
                    conn.commit()
            
            # 4. Registrar IPC aplicado
            self.db.execute_query(
                """INSERT INTO controlipc 
                (mes, anio, porcentajeipc, fechaaplicacion, aplicado, usuarioaplicacion) 
                VALUES (%s, %s, %s, %s, %s, %s)""",
                (datetime.now().strftime('%B'), datetime.now().year, 
                 porcentaje, datetime.now().date(), True, usuario)
            )
            
            # 5. Auditoría
            self.db.execute_query(
                "INSERT INTO audit_log (usuario, accion, detalles) VALUES (%s, %s, %s)",
                (usuario, 'IPC_APLICADO', f'IPC {porcentaje}% aplicado a {len(productos)} productos (LOTE)')
            )
            
            # Invalidar cache de productos para forzar recarga de nuevos precios
            self._invalidate_cache('productos')
            
            return True
            
        except Exception as e:
            logging.error(f"Error al aplicar IPC (LOTE): {e}. Trace: {traceback.format_exc()}")
            raise BusinessRuleError(f"No se pudo aplicar el IPC: {str(e)}")
    
    def get_historial_ipc(self):
        """Obtiene historial de IPC."""
        return self.db.execute_query(
            "SELECT * FROM controlipc ORDER BY anio DESC, mes DESC",
            fetch_all=True
        ) or []

    # ==================== MÉTODOS DE COMPROBANTES ====================
    
    @handle_db_errors
    def registrar_comprobante(self, tipo_doc, serie, numero, idproveedor, fecha_doc, total, detalles):
        """Registra un nuevo comprobante y sus detalles de productos."""
        try:
            # 1. Validación
            if not tipo_doc or not serie or not numero or not fecha_doc:
                raise ValidationError("Faltan campos obligatorios para el comprobante.")
            
            total_decimal = self.safe_decimal(total)
            
            # 2. Registrar Comprobante (Maestro)
            query_maestro = """
            INSERT INTO comprobantes 
            (tipo_doc, serie, numero, idproveedor, fecha_doc, total, fechacreacion, estado) 
            VALUES (%s, %s, %s, %s, %s, %s, NOW(), 'EMITIDO')
            """
            self.db.execute_query(query_maestro, (tipo_doc, serie, numero, idproveedor, fecha_doc, total_decimal))
            
            # 3. Obtener el ID del comprobante recién insertado
            comprobante = self.db.execute_query(
                "SELECT LAST_INSERT_ID() AS idcomprobante", 
                fetch_one=True
            )
            idcomprobante = comprobante['idcomprobante']
            
            # 4. Registrar Detalles (Transaccional)
            detalles_para_insert = []
            updates_stock = []
            
            for detalle in detalles:
                idproducto = self.safe_int(detalle.get('idproducto'))
                cantidad = self.safe_int(detalle.get('cantidad'))
                precio_unitario = self.safe_decimal(detalle.get('precio_unitario'))
                
                if idproducto == 0 or cantidad <= 0 or precio_unitario < 0:
                    raise ValidationError(f"Detalle de producto inválido: {detalle}")
                
                subtotal = cantidad * precio_unitario
                
                # (idcomprobante, idproducto, cantidad, precio_unitario, subtotal)
                detalles_para_insert.append((idcomprobante, idproducto, cantidad, precio_unitario, subtotal))
                
                # Preparar para actualizar stock: si es compra (COMPRA), se suma stock.
                if tipo_doc == 'COMPRA' or tipo_doc == 'NOTA_INGRESO':
                    updates_stock.append((cantidad, idproducto))
                elif tipo_doc == 'VENTA' or tipo_doc == 'NOTA_SALIDA':
                    # Para ventas, restar stock
                    updates_stock.append((-cantidad, idproducto))
            
            if detalles_para_insert:
                conn = self.db.connect()
                with conn.cursor() as cur:
                    insert_detalle_query = """
                    INSERT INTO detallecomprobantes 
                    (idcomprobante, idproducto, cantidad, precio_unitario, subtotal) 
                    VALUES (%s, %s, %s, %s, %s)
                    """
                    cur.executemany(insert_detalle_query, detalles_para_insert)
                    conn.commit()
            
            # 5. Actualizar Stock (si es necesario)
            if updates_stock:
                conn = self.db.connect()
                with conn.cursor() as cur:
                    update_stock_query = """
                    UPDATE productos 
                    SET stock = stock + %s 
                    WHERE idproducto = %s
                    """
                    cur.executemany(update_stock_query, updates_stock)
                    conn.commit()
                self._invalidate_cache('productos') # Invalidar cache de productos
            
            # 6. Auditoría
            self.db.execute_query(
                "INSERT INTO audit_log (usuario, accion, detalles) VALUES (%s, %s, %s)",
                ('system', 'COMPROBANTE_REGISTRADO', f'{tipo_doc} {serie}-{numero} con {len(detalles)} detalles.')
            )
            
            return idcomprobante
            
        except ValidationError as e:
            raise
        except Exception as e:
            logging.error(f"Error al registrar comprobante: {traceback.format_exc()}")
            raise BusinessRuleError(f"Error al registrar comprobante: {str(e)}")

    def get_comprobantes(self, force_refresh=False):
        """Obtiene todos los comprobantes con nombre de proveedor."""
        cache_key = 'comprobantes'
        if not force_refresh and self._is_cache_valid(cache_key):
            return self._comprobantes_cache
            
        query = """
        SELECT 
            c.idcomprobante, 
            c.tipo_doc, 
            c.serie, 
            c.numero, 
            p.nombre as nombreproveedor, 
            c.fecha_doc, 
            c.total, 
            c.estado,
            c.fechacreacion
        FROM comprobantes c
        LEFT JOIN proveedores p ON c.idproveedor = p.idproveedor
        ORDER BY c.fechacreacion DESC
        """
        self._comprobantes_cache = self.db.execute_query(query, fetch_all=True) or []
        self._update_cache_timestamp(cache_key)
        return self._comprobantes_cache
    
    def get_detalle_comprobante(self, idcomprobante):
        """Obtiene los detalles de un comprobante específico."""
        query = """
        SELECT 
            dc.iddetallecomprobante, 
            prod.nombre as nombreproducto, 
            dc.cantidad, 
            dc.precio_unitario, 
            dc.subtotal
        FROM detallecomprobantes dc
        JOIN productos prod ON dc.idproducto = prod.idproducto
        WHERE dc.idcomprobante = %s
        """
        return self.db.execute_query(query, (idcomprobante,), fetch_all=True) or []

    def get_comprobante_by_id(self, idcomprobante):
        """Obtiene un comprobante específico por ID."""
        query = """
        SELECT 
            c.idcomprobante, 
            c.tipo_doc, 
            c.serie, 
            c.numero, 
            c.idproveedor,
            p.nombre as nombreproveedor, 
            c.fecha_doc, 
            c.total, 
            c.estado,
            c.fechacreacion
        FROM comprobantes c
        LEFT JOIN proveedores p ON c.idproveedor = p.idproveedor
        WHERE c.idcomprobante = %s
        """
        return self.db.execute_query(query, (idcomprobante,), fetch_one=True)

    # ==================== MÉTODOS DE CACHE INTELIGENTE ====================
    
    def _is_cache_valid(self, cache_key, max_age_minutes=5):
        """Verifica si el cache es válido."""
        if cache_key not in self._cache_timestamp:
            return False
        cache_age = datetime.now() - self._cache_timestamp[cache_key]
        return cache_age < timedelta(minutes=max_age_minutes)

    def _update_cache_timestamp(self, cache_key):
        """Actualiza el timestamp del cache."""
        self._cache_timestamp[cache_key] = datetime.now()

    def _invalidate_cache(self, cache_key):
        """Invalida el cache específico."""
        if cache_key == 'usuarios':
            self._usuarios_cache = None
        elif cache_key == 'proveedores':
            self._proveedores_cache = None
        elif cache_key == 'productos':
            self._productos_cache = None
        elif cache_key == 'comprobantes':
            self._comprobantes_cache = None
        
        if cache_key in self._cache_timestamp:
            del self._cache_timestamp[cache_key]

    # ==================== MÉTODOS DE REPORTES ====================
    
    def get_estadisticas_sistema(self):
        """Obtiene estadísticas del sistema."""
        try:
            stats = {}
            
            # Contar usuarios
            users_count = self.db.execute_query(
                "SELECT COUNT(*) as count FROM usuarios WHERE activo = true",
                fetch_one=True
            )
            stats['usuarios_activos'] = users_count['count'] if users_count else 0
            
            # Contar proveedores
            providers_count = self.db.execute_query(
                "SELECT COUNT(*) as count FROM proveedores WHERE activo = true",
                fetch_one=True
            )
            stats['proveedores_activos'] = providers_count['count'] if providers_count else 0
            
            # Contar productos
            products_count = self.db.execute_query(
                "SELECT COUNT(*) as count FROM productos WHERE activo = true",
                fetch_one=True
            )
            stats['productos_activos'] = products_count['count'] if products_count else 0
            
            # Valor total del inventario
            inventory_value = self.db.execute_query(
                "SELECT SUM(stock * precio) as total FROM productos WHERE activo = true",
                fetch_one=True
            )
            stats['valor_inventario'] = inventory_value['total'] if inventory_value and inventory_value['total'] else Decimal('0.00')
            
            # Productos con stock bajo
            low_stock = self.db.execute_query(
                "SELECT COUNT(*) as count FROM productos WHERE stock < stockminimo AND activo = true",
                fetch_one=True
            )
            stats['productos_stock_bajo'] = low_stock['count'] if low_stock else 0
            
            # Comprobantes del mes
            month_comprobantes = self.db.execute_query(
                "SELECT COUNT(*) as count FROM comprobantes WHERE MONTH(fechacreacion) = MONTH(CURRENT_DATE()) AND YEAR(fechacreacion) = YEAR(CURRENT_DATE())",
                fetch_one=True
            )
            stats['comprobantes_mes'] = month_comprobantes['count'] if month_comprobantes else 0
            
            return stats
            
        except Exception as e:
            logging.error(f"Error al obtener estadísticas: {e}")
            return {}
            
    def get_data_for_chart(self, chart_type='productos_por_categoria'):
        """Obtiene datos para gráficos."""
        if chart_type == 'productos_por_categoria':
            query = """
            SELECT categoria, COUNT(*) as count
            FROM productos
            WHERE activo = TRUE
            GROUP BY categoria
            ORDER BY count DESC
            LIMIT 5
            """
            data = self.db.execute_query(query, fetch_all=True)
            return {
                'labels': [d['categoria'] for d in data],
                'values': [d['count'] for d in data],
                'title': 'Top 5 Categorías de Productos Activos'
            }

        elif chart_type == 'valor_inventario_proveedor':
            query = """
            SELECT p.nombre, SUM(prod.stock * prod.precio) as total_valor
            FROM proveedores p
            JOIN productos prod ON p.idproveedor = prod.idproveedor
            WHERE prod.activo = TRUE
            GROUP BY p.nombre
            ORDER BY total_valor DESC
            LIMIT 5
            """
            data = self.db.execute_query(query, fetch_all=True)
            return {
                'labels': [d['nombre'] for d in data],
                'values': [float(d['total_valor']) for d in data],
                'title': 'Valor de Inventario por Proveedor (Top 5)'
            }

        return {'labels': [], 'values': [], 'title': 'Gráfico de Datos'}

    def export_to_excel(self, data, filename, headers):
        """Exporta datos a archivo Excel."""
        if not EXCEL_AVAILABLE:
            raise ImportError("openpyxl no está instalado. Instale con: pip install openpyxl")

        wb = Workbook()
        ws = wb.active
        ws.title = "Datos Exportados"

        # Estilos
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="2563EB", end_color="2563EB", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")

        # Encabezados
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment

        # Datos
        for row_num, row_data in enumerate(data, 2):
            for col_num, value in enumerate(row_data, 1):
                ws.cell(row=row_num, column=col_num, value=value)

        # Autoajustar columnas
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2)
            ws.column_dimensions[column_letter].width = adjusted_width

        wb.save(filename)
        return filename

# ==============================================================================
# COMPONENTES UI MEJORADOS
# ==============================================================================

class AnimatedButton(QPushButton):
    """Botón con animaciones suaves con colores profesionales."""
    
    def __init__(self, text="", parent=None):
        super().__init__(text, parent)
        self._color = QColor(COLORS['primary'])
        self.update_style()
        
    def get_color(self):
        return self._color
        
    def set_color(self, color):
        self._color = color
        self.update_style()
        
    color = pyqtProperty(QColor, get_color, set_color)
    
    def update_style(self):
        self.setStyleSheet(f"""
            QPushButton {{
                background-color: {self._color.name()};
                color: white;
                border: none;
                border-radius: 6px;
                padding: 10px 20px;
                font-weight: 600;
                font-size: 13px;
                min-height: 16px;
            }}
            QPushButton:hover {{
                background-color: {self._color.lighter(110).name()};
                transform: translateY(-1px);
            }}
            QPushButton:pressed {{
                background-color: {self._color.darker(110).name()};
                transform: translateY(0px);
            }}
            QPushButton:disabled {{
                background-color: {COLORS['light']};
                color: {COLORS['text_secondary']};
                opacity: 0.6;
            }}
        """)

class ModernInput(QLineEdit):
    """Campo de entrada moderno con validación visual."""
    
    def __init__(self, placeholder="", parent=None):
        super().__init__(parent)
        self.setPlaceholderText(placeholder)
        self.setMinimumHeight(40)
        self.valid = True
        self.update_style()
        
    def set_valid(self, valid):
        self.valid = valid
        self.update_style()
        
    def update_style(self):
        border_color = COLORS['success'] if self.valid else COLORS['danger']
        self.setStyleSheet(f"""
            QLineEdit {{
                border: 1px solid {border_color};
                border-radius: 4px;
                padding: 8px 12px;
                font-size: 13px;
                background-color: {COLORS['background']};
                color: {COLORS['text_primary']};
                min-height: 16px;
            }}
            QLineEdit:focus {{
                border-color: {COLORS['accent']};
                background-color: {COLORS['background']};
                box-shadow: 0 0 0 2px {COLORS['shadow']};
            }}
            QLineEdit:disabled {{
                background-color: {COLORS['light']};
                color: {COLORS['text_secondary']};
                border-color: {COLORS['border']};
            }}
        """)

class CardWidget(QFrame):
    """Widget tipo tarjeta para agrupar contenido con estilo profesional."""
    
    def __init__(self, parent=None):
        super().__init__(parent)
        self.setFrameStyle(QFrame.StyledPanel)
        self.setStyleSheet(f"""
            QFrame {{
                background-color: {COLORS['background']};
                border-radius: 8px;
                border: 1px solid {COLORS['border']};
                padding: 0px;
                box-shadow: 0 1px 3px {COLORS['shadow']};
            }}
        """)
        self.setAttribute(Qt.WA_StyledBackground, True)

class HeaderWidget(QWidget):
    """Encabezado moderno para la aplicación con gradiente profesional."""
    
    def __init__(self, title, subtitle="", parent=None):
        super().__init__(parent)
        self.setFixedHeight(70)
        self.setStyleSheet(f"""
            background: qlineargradient(x1:0, y1:0, x2:1, y2:0,
                                       stop:0 {COLORS['primary']},
                                       stop:1 {COLORS['secondary']});
            border-bottom: 2px solid {COLORS['accent']};
        """)

        layout = QHBoxLayout()
        layout.setContentsMargins(25, 15, 25, 15)

        # Título
        title_layout = QVBoxLayout()
        self.title_label = QLabel(title)
        self.title_label.setStyleSheet("color: white; font-size: 20px; font-weight: 600; margin: 0;")
        self.subtitle_label = QLabel(subtitle)
        self.subtitle_label.setStyleSheet("color: rgba(255,255,255,0.8); font-size: 12px; margin: 0;")

        title_layout.addWidget(self.title_label)
        title_layout.addWidget(self.subtitle_label)

        layout.addLayout(title_layout)
        layout.addStretch()

        # Información de usuario
        self.user_info = QLabel()
        self.user_info.setStyleSheet(f"""
            color: white;
            font-size: 12px;
            background: rgba(255,255,255,0.15);
            padding: 6px 12px;
            border-radius: 4px;
            border: 1px solid rgba(255,255,255,0.2);
        """)
        layout.addWidget(self.user_info)

        self.setLayout(layout)
    
    def set_user_info(self, username, role):
        self.user_info.setText(f"👤 {username} | {role.upper()}")

# ==============================================================================
# DIÁLOGOS ESPECIALIZADOS
# ==============================================================================

class IPCDialog(QDialog):
    """Diálogo para aplicar variaciones de IPC."""
    
    def __init__(self, business_logic, username, parent=None):
        super().__init__(parent)
        self.business_logic = business_logic
        self.username = username
        self.setup_ui()
    
    def setup_ui(self):
        self.setWindowTitle("📈 Ajuste de Precios por IPC")
        self.setFixedSize(500, 400)
        self.setModal(True)
        
        layout = QVBoxLayout()
        layout.setContentsMargins(25, 25, 25, 25)
        
        # Header
        header = QLabel("Ajuste de Precios por Índice de Precios al Consumidor")
        header.setStyleSheet(f"""
            color: {COLORS['primary']};
            font-size: 18px;
            font-weight: bold;
            margin-bottom: 20px;
        """)
        layout.addWidget(header)
        
        # Información
        info_text = QLabel(
            "Esta herramienta aplicará un ajuste porcentual a todos los precios "
            "de productos activos en el sistema. El cambio se registrará en el "
            "historial de precios para auditoría."
        )
        info_text.setWordWrap(True)
        info_text.setStyleSheet(f"""
            background-color: {COLORS['light']};
            padding: 15px;
            border-radius: 8px;
            color: {COLORS['text_secondary']};
            margin-bottom: 20px;
        """)
        layout.addWidget(info_text)
        
        # Formulario
        form_layout = QFormLayout()
        form_layout.setLabelAlignment(Qt.AlignRight)
        form_layout.setVerticalSpacing(15)
        
        self.mes_combo = QComboBox()
        self.mes_combo.addItems(['Enero', 'Febrero', 'Marzo', 'Abril', 'Mayo', 'Junio',
                               'Julio', 'Agosto', 'Septiembre', 'Octubre', 'Noviembre', 'Diciembre'])
        self.mes_combo.setCurrentIndex(datetime.now().month - 1)
            
        self.anio_input = ModernInput()
        self.anio_input.setText(str(datetime.now().year))
        self.anio_input.setValidator(QIntValidator(2020, 2030))
        
        self.porcentaje_input = ModernInput()
        self.porcentaje_input.setValidator(QDoubleValidator(0.1, 100.0, 2))
        self.porcentaje_input.setPlaceholderText("Ej: 2.5")
        
        form_layout.addRow("Mes:", self.mes_combo)
        form_layout.addRow("Año:", self.anio_input)
        form_layout.addRow("Porcentaje IPC (%):", self.porcentaje_input)
        
        layout.addLayout(form_layout)
        
        # Historial IPC reciente
        historial_label = QLabel("Historial Reciente de IPC:")
        historial_label.setStyleSheet(f"color: {COLORS['text_primary']}; font-weight: bold; margin-top: 20px;")
        layout.addWidget(historial_label)
        
        self.historial_text = QTextEdit()
        self.historial_text.setMaximumHeight(100)
        self.historial_text.setReadOnly(True)
        self.historial_text.setStyleSheet(f"""
            background-color: {COLORS['surface']};
            border: 1px solid {COLORS['light']};
            border-radius: 6px;
            padding: 10px;
        """)
        layout.addWidget(self.historial_text)
        
        # Botones
        button_layout = QHBoxLayout()
        self.aplicar_btn = AnimatedButton("📈 Aplicar IPC")
        self.aplicar_btn.setMinimumHeight(45)
        self.aplicar_btn.clicked.connect(self.aplicar_ipc)
        
        cancelar_btn = AnimatedButton("❌ Cancelar")
        cancelar_btn.set_color(QColor(COLORS['danger']))
        cancelar_btn.clicked.connect(self.reject)
        
        button_layout.addWidget(self.aplicar_btn)
        button_layout.addStretch()
        button_layout.addWidget(cancelar_btn)
        
        layout.addLayout(button_layout)
        self.setLayout(layout)
        self.cargar_historial_ipc()

    def cargar_historial_ipc(self):
        """Carga el historial reciente de IPC."""
        try:
            historial = self.business_logic.get_historial_ipc()
            if historial:
                texto = ""
                for ipc in historial[:3]: # Últimos 3
                    estado = "✅ Aplicado" if ipc['aplicado'] else "⏳ Pendiente"
                    texto += f"{ipc['mes']} {ipc['anio']}: {ipc['porcentajeipc']}% - {estado}\n"
                self.historial_text.setText(texto)
            else:
                self.historial_text.setText("No hay historial de IPC registrado.")
        except Exception as e:
            self.historial_text.setText(f"Error cargando historial: {str(e)}")

    def aplicar_ipc(self):
        """Aplica el porcentaje de IPC a los productos."""
        porcentaje_text = self.porcentaje_input.text().strip().replace(',', '.')
        
        if not porcentaje_text:
            QMessageBox.warning(self, "Error de Validación", "Por favor ingrese el porcentaje de IPC.")
            return

        try:
            porcentaje = float(porcentaje_text)
            if porcentaje <= 0:
                QMessageBox.warning(self, "Error de Validación", "El porcentaje debe ser mayor a 0.")
                return

            # Confirmación
            confirm = QMessageBox.question(
                self, "Confirmar Aplicación de IPC", 
                f"¿Está seguro de aplicar un {porcentaje}% de IPC a todos los productos?\n\n"
                f"Esta acción afectará todos los precios activos y no se puede deshacer.",
                QMessageBox.Yes | QMessageBox.No
            )
            
            if confirm == QMessageBox.Yes:
                self.aplicar_btn.setEnabled(False)
                self.aplicar_btn.setText("Aplicando...")

                # Aplicar IPC
                success = self.business_logic.aplicar_ipc(porcentaje, self.username)

                if success:
                    QMessageBox.information(
                        self, "IPC Aplicado", 
                        f"✅ Se aplicó exitosamente el {porcentaje}% de IPC a todos los productos.\n\n"
                        f"Los nuevos precios ya están activos en el sistema."
                    )
                    self.accept()
                else:
                    QMessageBox.critical(self, "Error", "No se pudo aplicar el IPC. Por favor intente nuevamente.")
        
        except ValueError:
            QMessageBox.warning(self, "Error de Formato", "Porcentaje inválido. Use números con punto decimal (ej: 2.5).")
        except BusinessRuleError as e:
            QMessageBox.critical(self, "Error de Negocio", f"No se pudo aplicar el IPC:\n\n{str(e)}")
        except Exception as e:
            QMessageBox.critical(self, "Error Inesperado", f"Error inesperado al aplicar IPC:\n\n{str(e)}")
        finally:
            self.aplicar_btn.setEnabled(True)
            self.aplicar_btn.setText("📈 Aplicar IPC")

# ==============================================================================
# PANEL DE GESTIÓN DE PROVEEDORES (MEJORADO)
# ==============================================================================

class ProveedoresManagementPanel(QWidget):
    """Panel para la gestión completa de proveedores."""
    
    def __init__(self, business_logic, user_role, parent=None):
        super().__init__(parent)
        self.business_logic = business_logic
        self.user_role = user_role
        self.current_proveedor_id = None
        self.setup_ui()
        self.load_proveedores()
        self.apply_permissions()

    def setup_ui(self):
        main_layout = QHBoxLayout(self)
        main_layout.setContentsMargins(10, 10, 10, 10)
        
        # Left side: Proveedores List (Table)
        self.proveedores_table = QTableWidget()
        self.proveedores_table.setColumnCount(9)
        self.proveedores_table.setHorizontalHeaderLabels([
            "ID", "Nombre", "Contacto", "Teléfono", "Email", 
            "Dirección", "RUC", "Registro", "Activo"
        ])
        self.proveedores_table.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeToContents)
        self.proveedores_table.horizontalHeader().setStretchLastSection(True)
        self.proveedores_table.setSelectionBehavior(QTableWidget.SelectRows)
        self.proveedores_table.setEditTriggers(QTableWidget.NoEditTriggers)
        self.proveedores_table.cellClicked.connect(self.select_proveedor)
        
        # Styling the table
        self.proveedores_table.setStyleSheet(f"""
            QTableWidget {{
                background-color: {COLORS['background']};
                border: 1px solid {COLORS['border']};
                border-radius: 6px;
                gridline-color: {COLORS['border']};
                alternate-background-color: {COLORS['surface']};
            }}
            QHeaderView::section {{
                background-color: {COLORS['primary']};
                color: white;
                padding: 8px;
                border: none;
                font-weight: 600;
                font-size: 12px;
            }}
            QTableWidget::item {{
                padding: 6px;
                border-bottom: 1px solid {COLORS['border']};
            }}
            QTableWidget::item:selected {{
                background-color: {COLORS['accent']};
                color: white;
            }}
        """)
        
        # Right side: Proveedor Details/Form
        self.form_widget = CardWidget()
        self.form_widget.setMinimumWidth(400)
        form_layout = QVBoxLayout(self.form_widget)
        
        header = QLabel("Detalles/Registro de Proveedor")
        header.setStyleSheet(f"font-size: 18px; font-weight: bold; color: {COLORS['primary']}; margin-bottom: 10px;")
        form_layout.addWidget(header)
        
        scroll_area = QScrollArea()
        scroll_area.setWidgetResizable(True)
        scroll_content = QWidget()
        self.form_layout = QFormLayout(scroll_content)
        self.form_layout.setLabelAlignment(Qt.AlignRight)
        
        # Form Fields
        self.id_input = QLineEdit()
        self.id_input.setReadOnly(True)
        self.nombre_input = ModernInput("Nombre del Proveedor")
        self.contacto_input = ModernInput("Persona de Contacto")
        self.telefono_input = ModernInput("Teléfono")
        self.email_input = ModernInput("Email")
        self.direccion_input = QTextEdit()
        self.direccion_input.setPlaceholderText("Dirección completa...")
        self.direccion_input.setMinimumHeight(60)
        self.ruc_input = ModernInput("RUC (11 dígitos)")
        # Use QRegExpValidator instead of QIntValidator to handle large numbers
        from PyQt5.QtGui import QRegExpValidator
        from PyQt5.QtCore import QRegExp
        ruc_validator = QRegExpValidator(QRegExp(r'^\d{0,11}$'))
        self.ruc_input.setValidator(ruc_validator)
        self.ruc_input.setMaxLength(11)
        
        self.activo_checkbox = QCheckBox("Proveedor Activo")
        self.activo_checkbox.setChecked(True)
        
        self.form_layout.addRow("ID:", self.id_input)
        self.form_layout.addRow("Nombre:*", self.nombre_input)
        self.form_layout.addRow("Contacto:", self.contacto_input)
        self.form_layout.addRow("Teléfono:", self.telefono_input)
        self.form_layout.addRow("Email:", self.email_input)
        self.form_layout.addRow("Dirección:", self.direccion_input)
        self.form_layout.addRow("RUC:*", self.ruc_input)
        self.form_layout.addRow("Estado:", self.activo_checkbox)
        
        scroll_area.setWidget(scroll_content)
        form_layout.addWidget(scroll_area)
        
        # Form Buttons
        button_layout = QHBoxLayout()
        self.new_btn = AnimatedButton("➕ Nuevo")
        self.save_btn = AnimatedButton("💾 Guardar")
        self.delete_btn = AnimatedButton("🗑️ Eliminar")
        self.export_excel_btn = AnimatedButton("📊 Exportar Excel")

        self.new_btn.clicked.connect(self.clear_form)
        self.save_btn.clicked.connect(self.save_proveedor)
        self.delete_btn.clicked.connect(self.delete_proveedor)
        self.export_excel_btn.clicked.connect(self.exportar_proveedores_excel)

        button_layout.addWidget(self.new_btn)
        button_layout.addWidget(self.save_btn)
        button_layout.addWidget(self.delete_btn)
        button_layout.addWidget(self.export_excel_btn)

        form_layout.addLayout(button_layout)
        
        main_layout.addWidget(self.proveedores_table)
        main_layout.addWidget(self.form_widget)

    def apply_permissions(self):
        """Ajusta la visibilidad y capacidad de edición según el rol."""
        is_admin_or_compras = self.user_role in ['admin', 'compras']
        
        self.save_btn.setEnabled(is_admin_or_compras)
        self.new_btn.setEnabled(is_admin_or_compras)
        self.delete_btn.setEnabled(self.user_role == 'admin')

        self.nombre_input.setReadOnly(not is_admin_or_compras)
        self.contacto_input.setReadOnly(not is_admin_or_compras)
        self.telefono_input.setReadOnly(not is_admin_or_compras)
        self.email_input.setReadOnly(not is_admin_or_compras)
        self.direccion_input.setReadOnly(not is_admin_or_compras)
        self.ruc_input.setReadOnly(not is_admin_or_compras)
        self.activo_checkbox.setEnabled(is_admin_or_compras)

    def load_proveedores(self, force_refresh=False):
        """Carga y actualiza la tabla de proveedores."""
        try:
            proveedores = self.business_logic.get_all_proveedores(force_refresh=force_refresh)
            self.proveedores_table.setRowCount(0)
            
            for row_num, p in enumerate(proveedores):
                self.proveedores_table.insertRow(row_num)
                
                # Columnas: 0=ID, 1=Nombre, 2=Contacto, 3=Teléfono, 4=Email, 5=Dirección, 6=RUC, 7=Registro, 8=Activo
                
                self.proveedores_table.setItem(row_num, 0, QTableWidgetItem(str(p['idproveedor'])))
                self.proveedores_table.setItem(row_num, 1, QTableWidgetItem(p['nombre']))
                self.proveedores_table.setItem(row_num, 2, QTableWidgetItem(p['contacto'] or 'N/A'))
                self.proveedores_table.setItem(row_num, 3, QTableWidgetItem(p['telefono'] or 'N/A'))
                self.proveedores_table.setItem(row_num, 4, QTableWidgetItem(p['email'] or 'N/A'))
                self.proveedores_table.setItem(row_num, 5, QTableWidgetItem(p['direccion'] or 'N/A'))
                self.proveedores_table.setItem(row_num, 6, QTableWidgetItem(p['ruc'] or 'N/A'))
                self.proveedores_table.setItem(row_num, 7, QTableWidgetItem(str(p['fecha_registro']).split(' ')[0]))
                
                activo_item = QTableWidgetItem("✅ Activo" if p['activo'] else "❌ Inactivo")
                if not p['activo']:
                    activo_item.setBackground(QColor(COLORS['danger']))
                    activo_item.setForeground(QColor(Qt.white))
                self.proveedores_table.setItem(row_num, 8, activo_item)
                
            self.proveedores_table.resizeColumnsToContents()

        except Exception as e:
            logging.error(f"Error al cargar proveedores: {e}")
            QMessageBox.critical(self, "Error de Carga", f"No se pudieron cargar los proveedores: {str(e)}")

    def clear_form(self):
        """Limpia el formulario para un nuevo registro."""
        self.current_proveedor_id = None
        self.id_input.clear()
        self.nombre_input.clear()
        self.contacto_input.clear()
        self.telefono_input.clear()
        self.email_input.clear()
        self.direccion_input.clear()
        self.ruc_input.clear()
        self.activo_checkbox.setChecked(True)
        self.nombre_input.setFocus()
        self.proveedores_table.clearSelection()

    def select_proveedor(self, row, column):
        """Carga los datos del proveedor seleccionado en el formulario."""
        try:
            self.current_proveedor_id = int(self.proveedores_table.item(row, 0).text())
            
            proveedor_data = self.business_logic.get_proveedor_by_id(self.current_proveedor_id)
            
            if proveedor_data:
                self.id_input.setText(str(proveedor_data['idproveedor']))
                self.nombre_input.setText(proveedor_data['nombre'])
                self.contacto_input.setText(proveedor_data['contacto'] or '')
                self.telefono_input.setText(proveedor_data['telefono'] or '')
                self.email_input.setText(proveedor_data['email'] or '')
                self.direccion_input.setText(proveedor_data['direccion'] or '')
                self.ruc_input.setText(proveedor_data['ruc'] or '')
                self.activo_checkbox.setChecked(proveedor_data['activo'])
            
        except Exception as e:
            logging.error(f"Error al seleccionar proveedor: {e}")
            self.clear_form()

    def save_proveedor(self):
        """Guarda o actualiza el proveedor."""
        if self.user_role not in ['admin', 'compras']:
            QMessageBox.warning(self, "Permiso Denegado", "No tiene permiso para guardar/actualizar proveedores.")
            return

        try:
            # 1. Recolección de datos
            id_proveedor = self.current_proveedor_id
            nombre = self.nombre_input.text().strip()
            contacto = self.contacto_input.text().strip()
            telefono = self.telefono_input.text().strip()
            email = self.email_input.text().strip()
            direccion = self.direccion_input.toPlainText().strip()
            ruc = self.ruc_input.text().strip()
            activo = self.activo_checkbox.isChecked()
            
            # 2. Validación de datos
            if not nombre or not ruc:
                QMessageBox.warning(self, "Error de Validación", "Los campos Nombre y RUC son obligatorios.")
                return
            
            if ruc and len(ruc) != 11:
                QMessageBox.warning(self, "Error de Validación", "El RUC debe tener exactamente 11 dígitos.")
                return

            if email and not self.business_logic.validate_email(email):
                QMessageBox.warning(self, "Error de Validación", "El formato del email es inválido.")
                return

            data = {
                'nombre': nombre,
                'contacto': contacto or None,
                'telefono': telefono or None,
                'email': email or None,
                'direccion': direccion or None,
                'ruc': ruc,
                'activo': activo
            }

            if id_proveedor:
                data['idproveedor'] = id_proveedor
                self.business_logic.save_proveedor(data)
                QMessageBox.information(self, "Éxito", f"Proveedor '{nombre}' actualizado correctamente.")
            else:
                self.business_logic.save_proveedor(data)
                QMessageBox.information(self, "Éxito", f"Proveedor '{nombre}' registrado correctamente.")
                
            # Recargar datos y limpiar formulario
            self.load_proveedores(force_refresh=True)
            self.clear_form()
            
        except ValidationError as e:
            QMessageBox.warning(self, "Error de Validación", str(e))
        except BusinessRuleError as e:
            QMessageBox.warning(self, "Error de Negocio", str(e))
        except Exception as e:
            logging.error(f"Error al guardar proveedor: {e}")
            QMessageBox.critical(self, "Error de Base de Datos", f"No se pudo guardar el proveedor: {str(e)}")

    def delete_proveedor(self):
        """Elimina (lógicamente) un proveedor."""
        if self.user_role != 'admin':
            QMessageBox.warning(self, "Permiso Denegado", "Solo los administradores pueden eliminar proveedores.")
            return

        if not self.current_proveedor_id:
            QMessageBox.warning(self, "Advertencia", "Seleccione un proveedor para eliminar.")
            return

        proveedor_nombre = self.nombre_input.text()
        confirm = QMessageBox.question(
            self, "Confirmar Eliminación",
            f"¿Está seguro de INACTIVAR el proveedor '{proveedor_nombre}'?",
            QMessageBox.Yes | QMessageBox.No
        )

        if confirm == QMessageBox.Yes:
            try:
                self.business_logic.delete_proveedor(self.current_proveedor_id)
                QMessageBox.information(self, "Éxito", f"Proveedor '{proveedor_nombre}' inactivado correctamente.")

                self.load_proveedores(force_refresh=True)
                self.clear_form()
            except BusinessRuleError as e:
                QMessageBox.warning(self, "Error de Negocio", str(e))
            except Exception as e:
                logging.error(f"Error al inactivar proveedor: {e}")
                QMessageBox.critical(self, "Error", f"No se pudo inactivar el proveedor: {str(e)}")

    def exportar_proveedores_excel(self):
        """Exporta la lista de proveedores a Excel."""
        try:
            proveedores = self.business_logic.get_all_proveedores(force_refresh=True)
            if not proveedores:
                QMessageBox.warning(self, "Advertencia", "No hay proveedores para exportar.")
                return

            # Preparar datos
            headers = ["ID", "Nombre", "Contacto", "Teléfono", "Email", "Dirección", "RUC", "Fecha Registro", "Activo"]
            data = []
            for p in proveedores:
                data.append([
                    p['idproveedor'],
                    p['nombre'],
                    p['contacto'] or 'N/A',
                    p['telefono'] or 'N/A',
                    p['email'] or 'N/A',
                    p['direccion'] or 'N/A',
                    p['ruc'] or 'N/A',
                    str(p['fecha_registro']).split(' ')[0],
                    "Sí" if p['activo'] else "No"
                ])

            # Generar nombre de archivo
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"proveedores_{timestamp}.xlsx"

            # Exportar
            self.business_logic.export_to_excel(data, filename, headers)
            QMessageBox.information(self, "Éxito", f"Proveedores exportados correctamente a {filename}")

        except ImportError as e:
            QMessageBox.critical(self, "Error", f"Funcionalidad no disponible: {str(e)}")
        except Exception as e:
            logging.error(f"Error al exportar proveedores a Excel: {e}")
            QMessageBox.critical(self, "Error", f"No se pudo exportar a Excel: {str(e)}")

# ==============================================================================
# PANEL DE COMPROBANTES (MEJORADO)
# ==============================================================================

class ComprobantesPanel(QWidget):
    """Panel para la gestión y visualización de comprobantes."""
    
    def __init__(self, business_logic, user_role, parent=None):
        super().__init__(parent)
        self.business_logic = business_logic
        self.user_role = user_role
        self.setup_ui()
        self.load_comprobantes()
        self.apply_permissions()

    def setup_ui(self):
        main_layout = QVBoxLayout(self)
        main_layout.setContentsMargins(10, 10, 10, 10)
        
        # Header con botones
        header_layout = QHBoxLayout()
        title = QLabel("📑 Gestión de Comprobantes")
        title.setStyleSheet(f"font-size: 18px; font-weight: bold; color: {COLORS['primary']};")

        self.nuevo_btn = AnimatedButton("➕ Nuevo Comprobante")
        self.ver_btn = AnimatedButton("👁️ Ver Detalles")
        self.imprimir_btn = AnimatedButton("🖨️ Imprimir")
        self.exportar_excel_btn = AnimatedButton("📊 Exportar Excel")
        self.recargar_btn = AnimatedButton("🔄 Recargar")

        self.nuevo_btn.clicked.connect(self.nuevo_comprobante)
        self.ver_btn.clicked.connect(self.ver_detalles)
        self.imprimir_btn.clicked.connect(self.imprimir_comprobante)
        self.exportar_excel_btn.clicked.connect(self.exportar_comprobantes_excel)
        self.recargar_btn.clicked.connect(self.load_comprobantes)

        header_layout.addWidget(title)
        header_layout.addStretch()
        header_layout.addWidget(self.nuevo_btn)
        header_layout.addWidget(self.ver_btn)
        header_layout.addWidget(self.imprimir_btn)
        header_layout.addWidget(self.exportar_excel_btn)
        header_layout.addWidget(self.recargar_btn)

        main_layout.addLayout(header_layout)
        
        # Tabla de comprobantes
        self.comprobantes_table = QTableWidget()
        self.comprobantes_table.setColumnCount(8)
        self.comprobantes_table.setHorizontalHeaderLabels([
            "ID", "Tipo", "Serie", "Número", "Proveedor", 
            "Fecha Doc.", "Total (S/.)", "Estado"
        ])
        self.comprobantes_table.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeToContents)
        self.comprobantes_table.horizontalHeader().setStretchLastSection(True)
        self.comprobantes_table.setSelectionBehavior(QTableWidget.SelectRows)
        self.comprobantes_table.setEditTriggers(QTableWidget.NoEditTriggers)
        self.comprobantes_table.cellDoubleClicked.connect(self.ver_detalles)
        
        # Styling the table
        self.comprobantes_table.setStyleSheet(f"""
            QTableWidget {{
                background-color: {COLORS['background']};
                border: 1px solid {COLORS['border']};
                border-radius: 6px;
                gridline-color: {COLORS['border']};
                alternate-background-color: {COLORS['surface']};
            }}
            QHeaderView::section {{
                background-color: {COLORS['primary']};
                color: white;
                padding: 8px;
                border: none;
                font-weight: 600;
                font-size: 12px;
            }}
            QTableWidget::item {{
                padding: 6px;
                border-bottom: 1px solid {COLORS['border']};
            }}
            QTableWidget::item:selected {{
                background-color: {COLORS['accent']};
                color: white;
            }}
        """)
        
        main_layout.addWidget(self.comprobantes_table)

    def apply_permissions(self):
        """Ajusta los permisos según el rol."""
        can_edit = self.user_role in ['admin', 'compras']
        self.nuevo_btn.setEnabled(can_edit)

    def load_comprobantes(self, force_refresh=False):
        """Carga y actualiza la tabla de comprobantes."""
        try:
            comprobantes = self.business_logic.get_comprobantes(force_refresh=force_refresh)
            self.comprobantes_table.setRowCount(0)
            
            for row_num, c in enumerate(comprobantes):
                self.comprobantes_table.insertRow(row_num)
                
                # Columnas: 0=ID, 1=Tipo, 2=Serie, 3=Numero, 4=Proveedor, 5=Fecha, 6=Total, 7=Estado
                
                self.comprobantes_table.setItem(row_num, 0, QTableWidgetItem(str(c['idcomprobante'])))
                self.comprobantes_table.setItem(row_num, 1, QTableWidgetItem(c['tipo_doc']))
                self.comprobantes_table.setItem(row_num, 2, QTableWidgetItem(c['serie']))
                self.comprobantes_table.setItem(row_num, 3, QTableWidgetItem(c['numero']))
                self.comprobantes_table.setItem(row_num, 4, QTableWidgetItem(c['nombreproveedor'] or 'N/A'))
                self.comprobantes_table.setItem(row_num, 5, QTableWidgetItem(str(c['fecha_doc'])))
                self.comprobantes_table.setItem(row_num, 6, QTableWidgetItem(f"S/. {c['total']:,.2f}"))
                
                estado_item = QTableWidgetItem(c['estado'])
                if c['estado'] == 'EMITIDO':
                    estado_item.setBackground(QColor(COLORS['success']))
                    estado_item.setForeground(QColor(Qt.white))
                elif c['estado'] == 'ANULADO':
                    estado_item.setBackground(QColor(COLORS['danger']))
                    estado_item.setForeground(QColor(Qt.white))
                self.comprobantes_table.setItem(row_num, 7, estado_item)
                
            self.comprobantes_table.resizeColumnsToContents()

        except Exception as e:
            logging.error(f"Error al cargar comprobantes: {e}")
            QMessageBox.critical(self, "Error de Carga", f"No se pudieron cargar los comprobantes: {str(e)}")

    def get_selected_comprobante_id(self):
        """Obtiene el ID del comprobante seleccionado."""
        selected_items = self.comprobantes_table.selectedItems()
        if not selected_items:
            return None
        return int(self.comprobantes_table.item(selected_items[0].row(), 0).text())

    def nuevo_comprobante(self):
        """Abre el diálogo para crear un nuevo comprobante."""
        dialog = ComprobanteDialog(self.business_logic, self.user_role, self)
        if dialog.exec_() == QDialog.Accepted:
            self.load_comprobantes(force_refresh=True)

    def ver_detalles(self):
        """Muestra los detalles del comprobante seleccionado."""
        comprobante_id = self.get_selected_comprobante_id()
        if not comprobante_id:
            QMessageBox.warning(self, "Advertencia", "Seleccione un comprobante para ver sus detalles.")
            return
            
        dialog = DetalleComprobanteDialog(self.business_logic, comprobante_id, self)
        dialog.exec_()

    def imprimir_comprobante(self):
        """Imprime el comprobante seleccionado."""
        comprobante_id = self.get_selected_comprobante_id()
        if not comprobante_id:
            QMessageBox.warning(self, "Advertencia", "Seleccione un comprobante para imprimir.")
            return

        try:
            comprobante = self.business_logic.get_comprobante_by_id(comprobante_id)
            detalles = self.business_logic.get_detalle_comprobante(comprobante_id)

            # Crear documento de impresión
            printer = QPrinter(QPrinter.HighResolution)
            dialog = QPrintDialog(printer, self)
            if dialog.exec_() == QPrintDialog.Accepted:
                document = QTextDocument()
                html = self.generar_html_comprobante(comprobante, detalles)
                document.setHtml(html)
                document.print_(printer)
                QMessageBox.information(self, "Éxito", "Comprobante impreso correctamente.")

        except Exception as e:
            logging.error(f"Error al imprimir comprobante: {e}")
            QMessageBox.critical(self, "Error", f"No se pudo imprimir el comprobante: {str(e)}")

    def exportar_comprobantes_excel(self):
        """Exporta la lista de comprobantes a Excel."""
        try:
            comprobantes = self.business_logic.get_comprobantes(force_refresh=True)
            if not comprobantes:
                QMessageBox.warning(self, "Advertencia", "No hay comprobantes para exportar.")
                return

            # Preparar datos
            headers = ["ID", "Tipo", "Serie", "Número", "Proveedor", "Fecha Doc.", "Total (S/.)", "Estado"]
            data = []
            for c in comprobantes:
                data.append([
                    c['idcomprobante'],
                    c['tipo_doc'],
                    c['serie'],
                    c['numero'],
                    c['nombreproveedor'] or 'N/A',
                    str(c['fecha_doc']),
                    float(c['total']),
                    c['estado']
                ])

            # Generar nombre de archivo
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"comprobantes_{timestamp}.xlsx"

            # Exportar
            self.business_logic.export_to_excel(data, filename, headers)
            QMessageBox.information(self, "Éxito", f"Comprobantes exportados correctamente a {filename}")

        except ImportError as e:
            QMessageBox.critical(self, "Error", f"Funcionalidad no disponible: {str(e)}")
        except Exception as e:
            logging.error(f"Error al exportar comprobantes a Excel: {e}")
            QMessageBox.critical(self, "Error", f"No se pudo exportar a Excel: {str(e)}")

    def generar_html_comprobante(self, comprobante, detalles):
        """Genera HTML para la impresión del comprobante."""
        html = f"""
        <html>
        <head>
        <style>
            body {{ font-family: Arial, sans-serif; margin: 20px; }}
            .header {{ text-align: center; border-bottom: 2px solid #000; padding-bottom: 10px; margin-bottom: 20px; }}
            .empresa {{ font-size: 18px; font-weight: bold; }}
            .comprobante {{ font-size: 16px; }}
            .detalles {{ width: 100%; border-collapse: collapse; margin: 20px 0; }}
            .detalles th, .detalles td {{ border: 1px solid #000; padding: 8px; text-align: left; }}
            .detalles th {{ background-color: #f0f0f0; }}
            .total {{ text-align: right; font-weight: bold; font-size: 16px; margin-top: 20px; }}
            .footer {{ margin-top: 30px; text-align: center; font-size: 12px; color: #666; }}
        </style>
        </head>
        <body>
            <div class="header">
                <div class="empresa">SISTEMA DE GESTIÓN DE ALIMENTOS PRO</div>
                <div class="comprobante">{comprobante['tipo_doc']} - {comprobante['serie']}-{comprobante['numero']}</div>
            </div>
            
            <div>
                <p><strong>Proveedor:</strong> {comprobante['nombreproveedor'] or 'N/A'}</p>
                <p><strong>Fecha Documento:</strong> {comprobante['fecha_doc']}</p>
                <p><strong>Estado:</strong> {comprobante['estado']}</p>
            </div>
            
            <table class="detalles">
                <thead>
                    <tr>
                        <th>Producto</th>
                        <th>Cantidad</th>
                        <th>Precio Unit.</th>
                        <th>Subtotal</th>
                    </tr>
                </thead>
                <tbody>
        """
        
        for detalle in detalles:
            html += f"""
                    <tr>
                        <td>{detalle['nombreproducto']}</td>
                        <td>{detalle['cantidad']}</td>
                        <td>S/. {detalle['precio_unitario']:,.2f}</td>
                        <td>S/. {detalle['subtotal']:,.2f}</td>
                    </tr>
            """
        
        html += f"""
                </tbody>
            </table>
            
            <div class="total">
                <strong>TOTAL: S/. {comprobante['total']:,.2f}</strong>
            </div>
            
            <div class="footer">
                <p>Comprobante generado el {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}</p>
                <p>Sistema de Gestión de Alimentos Pro v{APP_VERSION}</p>
            </div>
        </body>
        </html>
        """
        
        return html

# ==============================================================================
# DIÁLOGO DE COMPROBANTE (MEJORADO)
# ==============================================================================

class ComprobanteDialog(QDialog):
    """Diálogo para crear/editar comprobantes."""
    
    def __init__(self, business_logic, user_role, parent=None):
        super().__init__(parent)
        self.business_logic = business_logic
        self.user_role = user_role
        self.detalles = []
        self.setup_ui()

    def setup_ui(self):
        self.setWindowTitle("📑 Nuevo Comprobante")
        self.setMinimumSize(800, 600)
        self.setModal(True)
        
        layout = QVBoxLayout(self)
        
        # Header
        header = QLabel("Registro de Comprobante")
        header.setStyleSheet(f"font-size: 18px; font-weight: bold; color: {COLORS['primary']}; margin-bottom: 10px;")
        layout.addWidget(header)
        
        # Formulario principal
        form_widget = CardWidget()
        form_layout = QFormLayout(form_widget)
        
        self.tipo_combo = QComboBox()
        self.tipo_combo.addItems(['COMPRA', 'VENTA', 'NOTA_INGRESO', 'NOTA_SALIDA'])
        
        self.serie_input = ModernInput()
        self.serie_input.setPlaceholderText("Ej: F001")
        
        self.numero_input = ModernInput()
        self.numero_input.setPlaceholderText("Ej: 00000001")
        
        self.proveedor_combo = QComboBox()
        self.load_proveedores_combo()
        
        self.fecha_input = QDateEdit()
        self.fecha_input.setDate(QDate.currentDate())
        self.fecha_input.setCalendarPopup(True)
        
        form_layout.addRow("Tipo Documento:*", self.tipo_combo)
        form_layout.addRow("Serie:*", self.serie_input)
        form_layout.addRow("Número:*", self.numero_input)
        form_layout.addRow("Proveedor:*", self.proveedor_combo)
        form_layout.addRow("Fecha Documento:*", self.fecha_input)
        
        layout.addWidget(form_widget)
        
        # Sección de detalles
        detalles_label = QLabel("Detalles del Comprobante")
        detalles_label.setStyleSheet(f"font-size: 16px; font-weight: bold; color: {COLORS['primary']}; margin-top: 20px;")
        layout.addWidget(detalles_label)
        
        # Controles para agregar detalles
        detalle_widget = CardWidget()
        detalle_layout = QHBoxLayout(detalle_widget)
        
        self.producto_combo = QComboBox()
        self.load_productos_combo()
        
        self.cantidad_input = QSpinBox()
        self.cantidad_input.setRange(1, 9999)
        self.cantidad_input.setValue(1)
        
        self.precio_input = QDoubleSpinBox()
        self.precio_input.setRange(0.01, 9999.99)
        self.precio_input.setDecimals(2)
        self.precio_input.setPrefix("S/. ")
        
        self.agregar_btn = AnimatedButton("➕ Agregar")
        self.agregar_btn.clicked.connect(self.agregar_detalle)
        
        detalle_layout.addWidget(QLabel("Producto:"))
        detalle_layout.addWidget(self.producto_combo)
        detalle_layout.addWidget(QLabel("Cantidad:"))
        detalle_layout.addWidget(self.cantidad_input)
        detalle_layout.addWidget(QLabel("Precio:"))
        detalle_layout.addWidget(self.precio_input)
        detalle_layout.addWidget(self.agregar_btn)
        
        layout.addWidget(detalle_widget)
        
        # Tabla de detalles
        self.detalles_table = QTableWidget()
        self.detalles_table.setColumnCount(5)
        self.detalles_table.setHorizontalHeaderLabels(["Producto", "Cantidad", "Precio Unit.", "Subtotal", "Acciones"])
        self.detalles_table.horizontalHeader().setStretchLastSection(True)
        
        layout.addWidget(self.detalles_table)
        
        # Total
        self.total_label = QLabel("Total: S/. 0.00")
        self.total_label.setStyleSheet(f"font-size: 16px; font-weight: bold; color: {COLORS['primary']};")
        layout.addWidget(self.total_label)
        
        # Botones
        button_layout = QHBoxLayout()
        self.guardar_btn = AnimatedButton("💾 Guardar Comprobante")
        self.cancelar_btn = AnimatedButton("❌ Cancelar")
        
        self.guardar_btn.clicked.connect(self.guardar_comprobante)
        self.cancelar_btn.clicked.connect(self.reject)
        
        button_layout.addWidget(self.guardar_btn)
        button_layout.addStretch()
        button_layout.addWidget(self.cancelar_btn)
        
        layout.addLayout(button_layout)

    def load_proveedores_combo(self):
        """Carga los proveedores en el combo."""
        try:
            proveedores = self.business_logic.get_all_proveedores()
            self.proveedor_combo.clear()
            self.proveedor_combo.addItem("Seleccionar Proveedor", userData=None)
            for p in proveedores:
                self.proveedor_combo.addItem(p['nombre'], userData=p['idproveedor'])
        except Exception as e:
            logging.error(f"Error al cargar proveedores: {e}")

    def load_productos_combo(self):
        """Carga los productos en el combo."""
        try:
            productos = self.business_logic.get_all_productos()
            self.producto_combo.clear()
            self.producto_combo.addItem("Seleccionar Producto", userData=None)
            for p in productos:
                self.producto_combo.addItem(p['nombre'], userData=p['idproducto'])
        except Exception as e:
            logging.error(f"Error al cargar productos: {e}")

    def agregar_detalle(self):
        """Agrega un detalle a la tabla."""
        try:
            producto_index = self.producto_combo.currentIndex()
            if producto_index == 0:
                QMessageBox.warning(self, "Error", "Seleccione un producto.")
                return
                
            idproducto = self.producto_combo.currentData()
            producto_nombre = self.producto_combo.currentText()
            cantidad = self.cantidad_input.value()
            precio = self.precio_input.value()
            subtotal = cantidad * precio
            
            # Agregar a la lista interna
            self.detalles.append({
                'idproducto': idproducto,
                'nombreproducto': producto_nombre,
                'cantidad': cantidad,
                'precio_unitario': precio,
                'subtotal': subtotal
            })
            
            # Actualizar tabla
            self.actualizar_tabla_detalles()
            self.actualizar_total()
            
            # Limpiar controles
            self.cantidad_input.setValue(1)
            self.precio_input.setValue(0.00)
            
        except Exception as e:
            logging.error(f"Error al agregar detalle: {e}")
            QMessageBox.critical(self, "Error", f"No se pudo agregar el detalle: {str(e)}")

    def eliminar_detalle(self, row):
        """Elimina un detalle de la tabla."""
        if 0 <= row < len(self.detalles):
            self.detalles.pop(row)
            self.actualizar_tabla_detalles()
            self.actualizar_total()

    def actualizar_tabla_detalles(self):
        """Actualiza la tabla de detalles."""
        self.detalles_table.setRowCount(len(self.detalles))
        
        for row, detalle in enumerate(self.detalles):
            self.detalles_table.setItem(row, 0, QTableWidgetItem(detalle['nombreproducto']))
            self.detalles_table.setItem(row, 1, QTableWidgetItem(str(detalle['cantidad'])))
            self.detalles_table.setItem(row, 2, QTableWidgetItem(f"S/. {detalle['precio_unitario']:,.2f}"))
            self.detalles_table.setItem(row, 3, QTableWidgetItem(f"S/. {detalle['subtotal']:,.2f}"))
            
            # Botón eliminar
            eliminar_btn = QPushButton("🗑️")
            eliminar_btn.setStyleSheet("QPushButton { background-color: #dc3545; color: white; border: none; padding: 5px; border-radius: 3px; }")
            eliminar_btn.clicked.connect(lambda checked, r=row: self.eliminar_detalle(r))
            self.detalles_table.setCellWidget(row, 4, eliminar_btn)
        
        self.detalles_table.resizeColumnsToContents()

    def actualizar_total(self):
        """Actualiza el total del comprobante."""
        total = sum(detalle['subtotal'] for detalle in self.detalles)
        self.total_label.setText(f"Total: S/. {total:,.2f}")

    def guardar_comprobante(self):
        """Guarda el comprobante en la base de datos."""
        try:
            # Validaciones
            if not self.serie_input.text().strip():
                QMessageBox.warning(self, "Error", "La serie es obligatoria.")
                return
                
            if not self.numero_input.text().strip():
                QMessageBox.warning(self, "Error", "El número es obligatorio.")
                return
                
            if self.proveedor_combo.currentIndex() == 0 and self.tipo_combo.currentText() in ['COMPRA', 'NOTA_INGRESO']:
                QMessageBox.warning(self, "Error", "Seleccione un proveedor para compras y notas de ingreso.")
                return
                
            if not self.detalles:
                QMessageBox.warning(self, "Error", "Agregue al menos un detalle al comprobante.")
                return

            # Preparar datos
            tipo_doc = self.tipo_combo.currentText()
            serie = self.serie_input.text().strip()
            numero = self.numero_input.text().strip()
            idproveedor = self.proveedor_combo.currentData() if self.proveedor_combo.currentIndex() > 0 else None
            fecha_doc = self.fecha_input.date().toString("yyyy-MM-dd")
            total = sum(detalle['subtotal'] for detalle in self.detalles)
            
            # Guardar comprobante
            comprobante_id = self.business_logic.registrar_comprobante(
                tipo_doc, serie, numero, idproveedor, fecha_doc, total, self.detalles
            )
            
            QMessageBox.information(
                self, "Éxito", 
                f"Comprobante {serie}-{numero} registrado correctamente.\n"
                f"ID: {comprobante_id}\n"
                f"Total: S/. {total:,.2f}"
            )
            
            self.accept()
            
        except ValidationError as e:
            QMessageBox.warning(self, "Error de Validación", str(e))
        except BusinessRuleError as e:
            QMessageBox.warning(self, "Error de Negocio", str(e))
        except Exception as e:
            logging.error(f"Error al guardar comprobante: {e}")
            QMessageBox.critical(self, "Error", f"No se pudo guardar el comprobante: {str(e)}")

# ==============================================================================
# DIÁLOGO DE DETALLE DE COMPROBANTE (MEJORADO)
# ==============================================================================

class DetalleComprobanteDialog(QDialog):
    """Diálogo para ver los detalles de un comprobante."""
    
    def __init__(self, business_logic, comprobante_id, parent=None):
        super().__init__(parent)
        self.business_logic = business_logic
        self.comprobante_id = comprobante_id
        self.setup_ui()
        self.cargar_datos()

    def setup_ui(self):
        self.setWindowTitle("👁️ Detalles del Comprobante")
        self.setFixedSize(600, 500)
        self.setModal(True)
        
        layout = QVBoxLayout(self)
        
        # Información del comprobante
        info_widget = CardWidget()
        info_layout = QFormLayout(info_widget)
        
        self.tipo_label = QLabel()
        self.serie_numero_label = QLabel()
        self.proveedor_label = QLabel()
        self.fecha_label = QLabel()
        self.total_label = QLabel()
        self.estado_label = QLabel()
        
        info_layout.addRow("Tipo Documento:", self.tipo_label)
        info_layout.addRow("Serie-Número:", self.serie_numero_label)
        info_layout.addRow("Proveedor:", self.proveedor_label)
        info_layout.addRow("Fecha Documento:", self.fecha_label)
        info_layout.addRow("Total:", self.total_label)
        info_layout.addRow("Estado:", self.estado_label)
        
        layout.addWidget(info_widget)
        
        # Detalles
        detalles_label = QLabel("Detalles del Comprobante")
        detalles_label.setStyleSheet(f"font-size: 16px; font-weight: bold; color: {COLORS['primary']}; margin-top: 10px;")
        layout.addWidget(detalles_label)
        
        self.detalles_table = QTableWidget()
        self.detalles_table.setColumnCount(4)
        self.detalles_table.setHorizontalHeaderLabels(["Producto", "Cantidad", "Precio Unit.", "Subtotal"])
        self.detalles_table.horizontalHeader().setStretchLastSection(True)
        self.detalles_table.setEditTriggers(QTableWidget.NoEditTriggers)
        
        layout.addWidget(self.detalles_table)
        
        # Botón cerrar
        cerrar_btn = AnimatedButton("Cerrar")
        cerrar_btn.clicked.connect(self.accept)
        layout.addWidget(cerrar_btn, alignment=Qt.AlignCenter)

    def cargar_datos(self):
        """Carga los datos del comprobante y sus detalles."""
        try:
            # Obtener comprobante
            comprobante = self.business_logic.get_comprobante_by_id(self.comprobante_id)
            if not comprobante:
                QMessageBox.critical(self, "Error", "No se encontró el comprobante.")
                self.reject()
                return
                
            # Mostrar información
            self.tipo_label.setText(comprobante['tipo_doc'])
            self.serie_numero_label.setText(f"{comprobante['serie']}-{comprobante['numero']}")
            self.proveedor_label.setText(comprobante['nombreproveedor'] or 'N/A')
            self.fecha_label.setText(str(comprobante['fecha_doc']))
            self.total_label.setText(f"S/. {comprobante['total']:,.2f}")
            self.estado_label.setText(comprobante['estado'])
            
            # Cargar detalles
            detalles = self.business_logic.get_detalle_comprobante(self.comprobante_id)
            self.detalles_table.setRowCount(len(detalles))
            
            for row, detalle in enumerate(detalles):
                self.detalles_table.setItem(row, 0, QTableWidgetItem(detalle['nombreproducto']))
                self.detalles_table.setItem(row, 1, QTableWidgetItem(str(detalle['cantidad'])))
                self.detalles_table.setItem(row, 2, QTableWidgetItem(f"S/. {detalle['precio_unitario']:,.2f}"))
                self.detalles_table.setItem(row, 3, QTableWidgetItem(f"S/. {detalle['subtotal']:,.2f}"))
            
            self.detalles_table.resizeColumnsToContents()
            
        except Exception as e:
            logging.error(f"Error al cargar detalles del comprobante: {e}")
            QMessageBox.critical(self, "Error", f"No se pudieron cargar los detalles: {str(e)}")

# ==============================================================================
# PANEL DE REPORTES Y GRÁFICOS (MEJORADO)
# ==============================================================================

class ChartCanvas(FigureCanvas):
    """Canvas para integrar gráficos Matplotlib en PyQt5."""
    def __init__(self, data, parent=None, width=5, height=4, dpi=100):
        self.fig, self.ax = plt.subplots(figsize=(width, height), dpi=dpi)
        super().__init__(self.fig)
        self.setParent(parent)
        self.plot_data(data)

    def plot_data(self, data):
        self.ax.clear()
        labels = data['labels']
        values = data['values']
        title = data['title']
        
        colors_list = [COLORS['primary'], COLORS['secondary'], COLORS['accent'], COLORS['success'], COLORS['warning']]
        
        # Gráfico de barras (o pie si hay pocos datos)
        if len(labels) <= 5:
            # Gráfico de pie para representación de distribución
            wedges, texts, autotexts = self.ax.pie(
                values, 
                labels=labels, 
                autopct='%1.1f%%', 
                startangle=90, 
                colors=colors_list[:len(labels)], 
                wedgeprops={'edgecolor': 'white', 'linewidth': 1}
            )
            self.ax.axis('equal') # Asegura que el pie sea circular
        else:
            # Gráfico de barras
            bars = self.ax.bar(labels, values, color=colors_list)
            self.ax.set_ylabel('Conteo/Valor')
            self.ax.set_xticklabels(labels, rotation=45, ha="right")
            
            # Mostrar valores en las barras
            for bar in bars:
                height = bar.get_height()
                self.ax.text(bar.get_x() + bar.get_width() / 2, height + 0.1, 
                             f'{height:g}', ha='center', va='bottom')

        self.ax.set_title(title, fontsize=14, fontweight='bold', color=COLORS['text_primary'])
        self.ax.patch.set_facecolor(COLORS['surface']) # Fondo del gráfico
        self.fig.tight_layout()
        self.draw()

class DashboardPanel(QWidget):
    """Panel de dashboard con estadísticas y gráficos."""
    
    def __init__(self, business_logic, parent=None):
        super().__init__(parent)
        self.business_logic = business_logic
        self.setup_ui()
        self.load_data()
        
    def setup_ui(self):
        main_layout = QVBoxLayout(self)
        main_layout.setSpacing(20)
        
        # Título principal
        title_label = QLabel("🚀 Dashboard Empresarial")
        title_label.setStyleSheet(f"font-size: 22px; font-weight: bold; color: {COLORS['primary']}; margin-top: 10px;")
        main_layout.addWidget(title_label)
        
        # 1. Panel de Estadísticas (Grid)
        self.stats_grid = QGridLayout()
        main_layout.addLayout(self.stats_grid)
        self.stat_widgets = {}
        
        stats_config = [
            ("Usuarios Activos", "usuarios_activos", "👥"),
            ("Proveedores Activos", "proveedores_activos", "🚚"),
            ("Productos Activos", "productos_activos", "📦"),
            ("Valor Inventario Total", "valor_inventario", "💰"),
            ("Stock Bajo", "productos_stock_bajo", "🚨"),
            ("Comprobantes (Mes)", "comprobantes_mes", "📑")
        ]
        
        row, col = 0, 0
        for title, key, icon in stats_config:
            card = self._create_stat_card(title, key, icon)
            self.stats_grid.addWidget(card, row, col)
            self.stat_widgets[key] = card.findChild(QLabel, f"{key}_value")
            
            col += 1
            if col > 2:
                col = 0
                row += 1
        
        # 2. Panel de Gráficos (Horizontal)
        charts_group = QGroupBox("Gráficos de Análisis Rápido")
        charts_group.setStyleSheet(f"QGroupBox {{ font-weight: bold; border: 1px solid {COLORS['light']}; border-radius: 8px; margin-top: 10px; }} QGroupBox::title {{ subcontrol-origin: margin; subcontrol-position: top left; padding: 0 10px; }}")
        charts_layout = QHBoxLayout(charts_group)
        
        # Gráfico 1: Productos por Categoría
        data1 = self.business_logic.get_data_for_chart('productos_por_categoria')
        self.chart1 = ChartCanvas(data1, self)
        charts_layout.addWidget(self.chart1)
        
        # Gráfico 2: Valor de Inventario por Proveedor
        data2 = self.business_logic.get_data_for_chart('valor_inventario_proveedor')
        self.chart2 = ChartCanvas(data2, self)
        charts_layout.addWidget(self.chart2)
        
        main_layout.addWidget(charts_group)
        
        main_layout.addStretch(1)
        
        # Botón de Recarga
        reload_btn = AnimatedButton("🔄 Recargar Datos")
        reload_btn.setMinimumWidth(200)
        reload_btn.set_color(QColor(COLORS['secondary']))
        reload_btn.clicked.connect(self.load_data)
        
        btn_layout = QHBoxLayout()
        btn_layout.addStretch()
        btn_layout.addWidget(reload_btn)
        btn_layout.addStretch()
        main_layout.addLayout(btn_layout)
        
    def _create_stat_card(self, title, key, icon):
        card = CardWidget()
        card_layout = QVBoxLayout(card)
        card_layout.setContentsMargins(15, 15, 15, 15)

        title_label = QLabel(f"{icon} {title}")
        title_label.setStyleSheet(f"font-size: 14px; font-weight: 500; color: {COLORS['text_secondary']};")

        value_label = QLabel("Cargando...")
        value_label.setObjectName(f"{key}_value")
        value_label.setStyleSheet(f"font-size: 30px; font-weight: bold; color: {COLORS['primary']}; margin-top: 5px;")

        card_layout.addWidget(title_label)
        card_layout.addWidget(value_label)
        card_layout.addStretch(1)

        card.setLayout(card_layout)
        return card

    def load_data(self):
        """Carga y actualiza todas las estadísticas y gráficos."""
        try:
            stats = self.business_logic.get_estadisticas_sistema()

            # Actualizar estadísticas
            self.stat_widgets['usuarios_activos'].setText(str(stats.get('usuarios_activos', 0)))
            self.stat_widgets['proveedores_activos'].setText(str(stats.get('proveedores_activos', 0)))
            self.stat_widgets['productos_activos'].setText(str(stats.get('productos_activos', 0)))

            valor_inventario = stats.get('valor_inventario', Decimal('0.00'))
            self.stat_widgets['valor_inventario'].setText(f"S/. {valor_inventario:,.2f}")

            low_stock_count = stats.get('productos_stock_bajo', 0)
            self.stat_widgets['productos_stock_bajo'].setText(str(low_stock_count))
            if low_stock_count > 0:
                self.stat_widgets['productos_stock_bajo'].setStyleSheet(f"font-size: 30px; font-weight: bold; color: {COLORS['danger']}; margin-top: 5px;")
            else:
                self.stat_widgets['productos_stock_bajo'].setStyleSheet(f"font-size: 30px; font-weight: bold; color: {COLORS['success']}; margin-top: 5px;")

            self.stat_widgets['comprobantes_mes'].setText(str(stats.get('comprobantes_mes', 0)))

            # Actualizar gráficos
            data1 = self.business_logic.get_data_for_chart('productos_por_categoria')
            self.chart1.plot_data(data1)

            data2 = self.business_logic.get_data_for_chart('valor_inventario_proveedor')
            self.chart2.plot_data(data2)

            # QMessageBox.information(self, "Datos Actualizados", "✅ El Dashboard se ha recargado exitosamente.")

        except Exception as e:
            logging.error(f"Error al cargar datos del dashboard: {e}")
            QMessageBox.critical(self, "Error de Carga", f"No se pudieron cargar los datos del Dashboard: {str(e)}")

# ==============================================================================
# PANEL DE GESTIÓN DE PRODUCTOS (MEJORADO)
# ==============================================================================

class ProductManagementPanel(QWidget):
    """Panel para la gestión de productos."""
    
    def __init__(self, business_logic, user_role, parent=None):
        super().__init__(parent)
        self.business_logic = business_logic
        self.user_role = user_role
        self.current_product_id = None
        self.setup_ui()
        self.load_productos()
        self.apply_permissions()

    def setup_ui(self):
        main_layout = QHBoxLayout(self)
        main_layout.setContentsMargins(10, 10, 10, 10)
        
        # Left side: Product List (Table)
        self.product_table = QTableWidget()
        self.product_table.setColumnCount(11)
        self.product_table.setHorizontalHeaderLabels([
            "ID", "Nombre", "Proveedor", "Categoría", "Stock", 
            "Stock Mín.", "Precio Venta (S/.)", "Precio Base (S/.)", 
            "Código Barras", "Registro", "Activo"
        ])
        self.product_table.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeToContents)
        self.product_table.horizontalHeader().setStretchLastSection(True)
        self.product_table.setSelectionBehavior(QTableWidget.SelectRows)
        self.product_table.setEditTriggers(QTableWidget.NoEditTriggers)
        self.product_table.cellClicked.connect(self.select_product)
        
        # Styling the table
        self.product_table.setStyleSheet(f"""
            QTableWidget {{
                background-color: {COLORS['background']};
                border: 1px solid {COLORS['border']};
                border-radius: 6px;
                gridline-color: {COLORS['border']};
                alternate-background-color: {COLORS['surface']};
            }}
            QHeaderView::section {{
                background-color: {COLORS['primary']};
                color: white;
                padding: 8px;
                border: none;
                font-weight: 600;
                font-size: 12px;
            }}
            QTableWidget::item {{
                padding: 6px;
                border-bottom: 1px solid {COLORS['border']};
            }}
            QTableWidget::item:selected {{
                background-color: {COLORS['accent']};
                color: white;
            }}
        """)
        
        # Right side: Product Details/Form
        self.form_widget = CardWidget()
        self.form_widget.setMinimumWidth(400)
        form_layout = QVBoxLayout(self.form_widget)
        
        header = QLabel("Detalles/Registro de Producto")
        header.setStyleSheet(f"font-size: 18px; font-weight: bold; color: {COLORS['primary']}; margin-bottom: 10px;")
        form_layout.addWidget(header)
        
        scroll_area = QScrollArea()
        scroll_area.setWidgetResizable(True)
        scroll_content = QWidget()
        self.form_layout = QFormLayout(scroll_content)
        self.form_layout.setLabelAlignment(Qt.AlignRight)
        
        # Form Fields
        self.id_input = QLineEdit()
        self.id_input.setReadOnly(True)
        self.nombre_input = ModernInput("Nombre del Producto")
        self.descripcion_input = QTextEdit()
        self.descripcion_input.setPlaceholderText("Descripción detallada...")
        self.descripcion_input.setMinimumHeight(60)
        
        self.categoria_combo = QComboBox()
        self.categoria_combo.addItems(["", "Frutas", "Verduras", "Carnes", "Pescados", "Lácteos", "Granos", "Bebidas", "Otros"])
        
        self.proveedor_combo = QComboBox()
        self.load_proveedores_combo()
        
        self.stock_input = QSpinBox()
        self.stock_input.setRange(0, 99999)
        self.stockminimo_input = QSpinBox()
        self.stockminimo_input.setRange(0, 99999)
        
        self.precio_input = QDoubleSpinBox()
        self.precio_input.setRange(0.01, 9999.99)
        self.precio_input.setDecimals(2)
        self.precio_input.setPrefix("S/. ")
        
        self.preciobase_input = QDoubleSpinBox()
        self.preciobase_input.setRange(0.01, 9999.99)
        self.preciobase_input.setDecimals(2)
        self.preciobase_input.setPrefix("S/. ")
        
        self.codigo_barras_input = ModernInput("Código de Barras (opcional)")
        
        self.activo_checkbox = QCheckBox("Producto Activo")
        self.activo_checkbox.setChecked(True)
        
        self.form_layout.addRow("ID:", self.id_input)
        self.form_layout.addRow("Nombre:*", self.nombre_input)
        self.form_layout.addRow("Descripción:", self.descripcion_input)
        self.form_layout.addRow("Categoría:*", self.categoria_combo)
        self.form_layout.addRow("Proveedor:*", self.proveedor_combo)
        self.form_layout.addRow("Stock:", self.stock_input)
        self.form_layout.addRow("Stock Mínimo:", self.stockminimo_input)
        self.form_layout.addRow("Precio Venta:*", self.precio_input)
        self.form_layout.addRow("Precio Base:", self.preciobase_input)
        self.form_layout.addRow("Cód. Barras:", self.codigo_barras_input)
        self.form_layout.addRow("Estado:", self.activo_checkbox)
        
        scroll_area.setWidget(scroll_content)
        form_layout.addWidget(scroll_area)
        
        # Form Buttons
        button_layout = QHBoxLayout()
        self.new_btn = AnimatedButton("➕ Nuevo")
        self.save_btn = AnimatedButton("💾 Guardar")
        self.delete_btn = AnimatedButton("🗑️ Eliminar")
        self.ipc_btn = AnimatedButton("📈 Ajuste IPC")
        self.export_excel_btn = AnimatedButton("📊 Exportar Excel")
        self.ipc_btn.set_color(QColor(COLORS['accent']))

        self.new_btn.clicked.connect(self.clear_form)
        self.save_btn.clicked.connect(self.save_product)
        self.delete_btn.clicked.connect(self.delete_product)
        self.ipc_btn.clicked.connect(self.show_ipc_dialog)
        self.export_excel_btn.clicked.connect(self.exportar_productos_excel)

        button_layout.addWidget(self.new_btn)
        button_layout.addWidget(self.save_btn)
        button_layout.addWidget(self.delete_btn)
        button_layout.addWidget(self.ipc_btn)
        button_layout.addWidget(self.export_excel_btn)

        form_layout.addLayout(button_layout)
        
        main_layout.addWidget(self.product_table)
        main_layout.addWidget(self.form_widget)

    def apply_permissions(self):
        """Ajusta la visibilidad y capacidad de edición según el rol."""
        is_admin_or_inv = self.user_role in ['admin', 'inventario', 'compras']
        is_admin_or_fin = self.user_role in ['admin', 'finanzas']
        
        self.save_btn.setEnabled(is_admin_or_inv)
        self.new_btn.setEnabled(is_admin_or_inv)
        self.delete_btn.setEnabled(self.user_role == 'admin')
        self.ipc_btn.setEnabled(is_admin_or_fin)

        self.nombre_input.setReadOnly(not is_admin_or_inv)
        self.descripcion_input.setReadOnly(not is_admin_or_inv)
        self.categoria_combo.setEnabled(is_admin_or_inv)
        self.proveedor_combo.setEnabled(is_admin_or_inv)
        self.stock_input.setReadOnly(not is_admin_or_inv)
        self.stockminimo_input.setReadOnly(not is_admin_or_inv)
        self.codigo_barras_input.setReadOnly(not is_admin_or_inv)
        self.activo_checkbox.setEnabled(is_admin_or_inv)

        # Finanzas puede ver pero no editar precios, solo aplicar IPC
        can_edit_prices = self.user_role == 'admin'
        self.precio_input.setReadOnly(not can_edit_prices)
        self.preciobase_input.setReadOnly(not can_edit_prices)

    def load_proveedores_combo(self):
        """Carga los proveedores en el QComboBox."""
        try:
            proveedores = self.business_logic.get_all_proveedores()
            self.proveedor_combo.clear()
            self.proveedor_combo.addItem("Seleccionar Proveedor", userData=None)
            for p in proveedores:
                self.proveedor_combo.addItem(p['nombre'], userData=p['idproveedor'])
                
        except Exception as e:
            QMessageBox.critical(self, "Error de Carga", f"No se pudo cargar la lista de proveedores: {str(e)}")

    def load_productos(self, force_refresh=False):
        """Carga y actualiza la tabla de productos."""
        try:
            productos = self.business_logic.get_all_productos(force_refresh=force_refresh)
            self.product_table.setRowCount(0)
            
            for row_num, p in enumerate(productos):
                self.product_table.insertRow(row_num)
                
                # Columnas: 0=ID, 1=Nombre, 2=Proveedor, 3=Categoria, 4=Stock, 5=StockMin, 6=Precio, 7=PrecioBase, 8=CodBarras, 9=Registro, 10=Activo
                
                # ID (Hidden from resize to contents, but kept for reference)
                id_item = QTableWidgetItem(str(p['idproducto']))
                id_item.setTextAlignment(Qt.AlignCenter)
                self.product_table.setItem(row_num, 0, id_item)
                
                # Nombre y otros
                self.product_table.setItem(row_num, 1, QTableWidgetItem(p['nombre']))
                self.product_table.setItem(row_num, 2, QTableWidgetItem(p['nombreproveedor']))
                self.product_table.setItem(row_num, 3, QTableWidgetItem(p['categoria']))
                
                # Stock (Color coding for low stock)
                stock_item = QTableWidgetItem(str(p['stock']))
                stock_item.setTextAlignment(Qt.AlignCenter)
                if p['stock'] < p['stockminimo']:
                    stock_item.setBackground(QColor(COLORS['danger']))
                    stock_item.setForeground(QColor(Qt.white))
                    stock_item.setToolTip(f"¡Stock bajo! Mínimo: {p['stockminimo']}")
                self.product_table.setItem(row_num, 4, stock_item)
                
                # Stock Mínimo
                stock_min_item = QTableWidgetItem(str(p['stockminimo']))
                stock_min_item.setTextAlignment(Qt.AlignCenter)
                self.product_table.setItem(row_num, 5, stock_min_item)
                
                # Precios
                self.product_table.setItem(row_num, 6, QTableWidgetItem(f"S/. {p['precio']:,.2f}"))
                self.product_table.setItem(row_num, 7, QTableWidgetItem(f"S/. {p['preciobase']:,.2f}"))
                
                # Otros campos
                self.product_table.setItem(row_num, 8, QTableWidgetItem(p['codigo_barras'] if p['codigo_barras'] else 'N/A'))
                self.product_table.setItem(row_num, 9, QTableWidgetItem(str(p['fecharegistro']).split(' ')[0]))
                self.product_table.setItem(row_num, 10, QTableWidgetItem("✅ Activo" if p['activo'] else "❌ Inactivo"))
                
            self.product_table.resizeColumnsToContents()

        except Exception as e:
            logging.error(f"Error al cargar productos: {e}")
            QMessageBox.critical(self, "Error de Carga", f"No se pudieron cargar los productos: {str(e)}")

    def clear_form(self):
        """Limpia el formulario para un nuevo registro."""
        self.current_product_id = None
        self.id_input.clear()
        self.nombre_input.clear()
        self.descripcion_input.clear()
        self.categoria_combo.setCurrentIndex(0)
        self.proveedor_combo.setCurrentIndex(0)
        self.stock_input.setValue(0)
        self.stockminimo_input.setValue(0)
        self.precio_input.setValue(0.00)
        self.preciobase_input.setValue(0.00)
        self.codigo_barras_input.clear()
        self.activo_checkbox.setChecked(True)
        self.nombre_input.setFocus()
        self.product_table.clearSelection()

    def select_product(self, row, column):
        """Carga los datos del producto seleccionado en el formulario."""
        try:
            self.current_product_id = int(self.product_table.item(row, 0).text())
            
            # Obtener el producto completo para cargar todos los campos
            all_productos = self.business_logic.get_all_productos()
            product_data = next((p for p in all_productos if p['idproducto'] == self.current_product_id), None)
            
            if product_data:
                self.id_input.setText(str(product_data['idproducto']))
                self.nombre_input.setText(product_data['nombre'])
                self.descripcion_input.setText(product_data['descripcion'])
                self.categoria_combo.setCurrentText(product_data['categoria'])
                
                # Encontrar el índice del proveedor por su nombre
                index = self.proveedor_combo.findText(product_data['nombreproveedor'])
                if index != -1:
                    self.proveedor_combo.setCurrentIndex(index)
                
                self.stock_input.setValue(product_data['stock'])
                self.stockminimo_input.setValue(product_data['stockminimo'])
                self.precio_input.setValue(float(product_data['precio']))
                self.preciobase_input.setValue(float(product_data['preciobase']))
                self.codigo_barras_input.setText(product_data['codigo_barras'] if product_data['codigo_barras'] else "")
                self.activo_checkbox.setChecked(product_data['activo'])
            
        except Exception as e:
            logging.error(f"Error al seleccionar producto: {e}")
            self.clear_form()

    def save_product(self):
        """Guarda o actualiza el producto."""
        if self.user_role not in ['admin', 'inventario', 'compras']:
            QMessageBox.warning(self, "Permiso Denegado", "No tiene permiso para guardar/actualizar productos.")
            return

        try:
            # 1. Recolección de datos
            id_producto = self.current_product_id
            nombre = self.nombre_input.text().strip()
            descripcion = self.descripcion_input.toPlainText().strip()
            categoria = self.categoria_combo.currentText()
            id_proveedor = self.proveedor_combo.currentData()
            stock = self.stock_input.value()
            stock_minimo = self.stockminimo_input.value()
            precio = self.precio_input.value()
            precio_base = self.preciobase_input.value()
            codigo_barras = self.codigo_barras_input.text().strip()
            activo = self.activo_checkbox.isChecked()
            
            # 2. Validación de datos
            if not nombre or not categoria or id_proveedor is None:
                QMessageBox.warning(self, "Error de Validación", "Los campos Nombre, Categoría y Proveedor son obligatorios.")
                return
            
            if precio <= 0.00:
                QMessageBox.warning(self, "Error de Validación", "El Precio de Venta debe ser mayor a cero.")
                return

            data = {
                'nombre': nombre,
                'descripcion': descripcion,
                'categoria': categoria,
                'idproveedor': id_proveedor,
                'stock': stock,
                'stockminimo': stock_minimo,
                'precio': self.business_logic.safe_decimal(precio),
                'preciobase': self.business_logic.safe_decimal(precio_base),
                'codigo_barras': codigo_barras,
                'activo': activo
            }

            if id_producto:
                # Actualizar
                query = """
                UPDATE productos SET 
                    nombre = %s, descripcion = %s, categoria = %s, idproveedor = %s, 
                    stock = %s, stockminimo = %s, precio = %s, preciobase = %s, 
                    codigo_barras = %s, activo = %s
                WHERE idproducto = %s
                """
                params = (
                    data['nombre'], data['descripcion'], data['categoria'], data['idproveedor'],
                    data['stock'], data['stockminimo'], data['precio'], data['preciobase'],
                    data['codigo_barras'], data['activo'], id_producto
                )
                self.business_logic.db.execute_query(query, params)
                QMessageBox.information(self, "Éxito", f"Producto '{nombre}' actualizado correctamente.")
                
            else:
                # Insertar Nuevo
                query = """
                INSERT INTO productos 
                (nombre, descripcion, categoria, idproveedor, stock, stockminimo, precio, preciobase, codigo_barras, activo, fecharegistro)
                VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, NOW())
                """
                params = (
                    data['nombre'], data['descripcion'], data['categoria'], data['idproveedor'],
                    data['stock'], data['stockminimo'], data['precio'], data['preciobase'],
                    data['codigo_barras'], data['activo']
                )
                self.business_logic.db.execute_query(query, params)
                QMessageBox.information(self, "Éxito", f"Producto '{nombre}' registrado correctamente.")
                
            # Recargar datos y limpiar formulario
            self.business_logic._invalidate_cache('productos')
            self.load_productos(force_refresh=True)
            self.clear_form()
            
        except ValidationError as e:
            QMessageBox.warning(self, "Error de Validación", str(e))
        except Exception as e:
            logging.error(f"Error al guardar producto: {e}")
            QMessageBox.critical(self, "Error de Base de Datos", f"No se pudo guardar el producto: {str(e)}")

    def delete_product(self):
        """Elimina (lógicamente) un producto."""
        if self.user_role != 'admin':
            QMessageBox.warning(self, "Permiso Denegado", "Solo los administradores pueden eliminar productos.")
            return

        if not self.current_product_id:
            QMessageBox.warning(self, "Advertencia", "Seleccione un producto para eliminar.")
            return
            
        confirm = QMessageBox.question(
            self, "Confirmar Eliminación", 
            f"¿Está seguro de INACTIVAR el producto ID {self.current_product_id}? (Se recomienda inactivar antes que eliminar permanentemente)",
            QMessageBox.Yes | QMessageBox.No
        )
        
        if confirm == QMessageBox.Yes:
            try:
                # Eliminación lógica (se cambia el estado a inactivo)
                query = "UPDATE productos SET activo = FALSE WHERE idproducto = %s"
                self.business_logic.db.execute_query(query, (self.current_product_id,))
                
                QMessageBox.information(self, "Éxito", f"Producto ID {self.current_product_id} inactivado correctamente.")
                
                self.business_logic._invalidate_cache('productos')
                self.load_productos(force_refresh=True)
                self.clear_form()
            except Exception as e:
                logging.error(f"Error al inactivar producto: {e}")
                QMessageBox.critical(self, "Error", f"No se pudo inactivar el producto: {str(e)}")

    def show_ipc_dialog(self):
        """Muestra el diálogo para aplicar IPC."""
        if self.user_role not in ['admin', 'finanzas']:
            QMessageBox.warning(self, "Permiso Denegado", "No tiene permiso para aplicar el ajuste IPC.")
            return

        dialog = IPCDialog(self.business_logic, self.parent().username, self)
        if dialog.exec_() == QDialog.Accepted:
            # Si el IPC se aplica, forzamos la recarga de la tabla de productos
            self.load_productos(force_refresh=True)

    def exportar_productos_excel(self):
        """Exporta la lista de productos a Excel."""
        try:
            productos = self.business_logic.get_all_productos(force_refresh=True)
            if not productos:
                QMessageBox.warning(self, "Advertencia", "No hay productos para exportar.")
                return

            # Preparar datos
            headers = ["ID", "Nombre", "Proveedor", "Categoría", "Stock", "Stock Mínimo", "Precio Venta (S/.)", "Precio Base (S/.)", "Código Barras", "Fecha Registro", "Activo"]
            data = []
            for p in productos:
                data.append([
                    p['idproducto'],
                    p['nombre'],
                    p['nombreproveedor'],
                    p['categoria'],
                    p['stock'],
                    p['stockminimo'],
                    float(p['precio']),
                    float(p['preciobase']),
                    p['codigo_barras'] or 'N/A',
                    str(p['fecharegistro']).split(' ')[0],
                    "Sí" if p['activo'] else "No"
                ])

            # Generar nombre de archivo
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"productos_{timestamp}.xlsx"

            # Exportar
            self.business_logic.export_to_excel(data, filename, headers)
            QMessageBox.information(self, "Éxito", f"Productos exportados correctamente a {filename}")

        except ImportError as e:
            QMessageBox.critical(self, "Error", f"Funcionalidad no disponible: {str(e)}")
        except Exception as e:
            logging.error(f"Error al exportar productos a Excel: {e}")
            QMessageBox.critical(self, "Error", f"No se pudo exportar a Excel: {str(e)}")

# ==============================================================================
# VENTANA PRINCIPAL COMPLETA (MEJORADA)
# ==============================================================================

class ModernMainWindow(QMainWindow):
    """Ventana principal de la aplicación, moderna y empresarial."""
    
    def __init__(self, business_logic, user_id, username, user_role, config_manager):
        super().__init__()
        self.business_logic = business_logic
        self.user_id = user_id
        self.username = username
        self.user_role = user_role
        self.config_manager = config_manager
        self.setup_ui()
        self.setup_tabs()
        self.show_welcome_message()

    def setup_ui(self):
        self.setWindowTitle(f"{APP_NAME} - {self.username.capitalize()} ({self.user_role.upper()})")
        self.setMinimumSize(1200, 800)
        
        # Establecer paleta de colores para un estilo unificado
        palette = QPalette()
        palette.setColor(QPalette.Window, QColor(COLORS['background']))
        palette.setColor(QPalette.WindowText, QColor(COLORS['text_primary']))
        palette.setColor(QPalette.Base, QColor(COLORS['surface']))
        palette.setColor(QPalette.AlternateBase, QColor(COLORS['light']))
        palette.setColor(QPalette.ToolTipBase, QColor(COLORS['dark']))
        palette.setColor(QPalette.ToolTipText, QColor(Qt.white))
        palette.setColor(QPalette.Text, QColor(COLORS['text_primary']))
        palette.setColor(QPalette.Button, QColor(COLORS['light']))
        palette.setColor(QPalette.ButtonText, QColor(COLORS['text_primary']))
        palette.setColor(QPalette.Highlight, QColor(COLORS['primary']))
        palette.setColor(QPalette.HighlightedText, QColor(Qt.white))
        self.setPalette(palette)
        
        self.setStyleSheet(f"""
            QMainWindow {{ background-color: {COLORS['background']}; }}
        """)
        
        central_widget = QWidget()
        main_layout = QVBoxLayout()
        main_layout.setContentsMargins(0, 0, 0, 0)
        
        # Header (Encabezado superior)
        self.header = HeaderWidget(APP_NAME, "Sistema Profesional de Gestión Alimentaria")
        self.header.set_user_info(self.username, self.user_role)
        main_layout.addWidget(self.header)
        
        # Tab Widget (Contenido principal)
        self.tab_widget = QTabWidget()
        self.tab_widget.setStyleSheet(f"""
            QTabWidget::pane {{ 
                border: 1px solid {COLORS['light']}; 
                border-top: 0px solid {COLORS['light']}; 
                background-color: {COLORS['background']}; 
            }}
            QTabBar::tab {{ 
                background: {COLORS['light']}; 
                border: 1px solid {COLORS['light']}; 
                border-bottom: none; 
                border-top-left-radius: 4px; 
                border-top-right-radius: 4px; 
                min-width: 100px; 
                padding: 10px; 
                color: {COLORS['text_secondary']}; 
                margin-right: 2px;
            }}
            QTabBar::tab:selected, QTabBar::tab:hover {{ 
                background: {COLORS['surface']}; 
                color: {COLORS['text_primary']}; 
            }}
            QTabBar::tab:selected {{ 
                font-weight: bold; 
                border-color: {COLORS['primary']}; 
                border-bottom-color: {COLORS['surface']}; 
            }}
        """)
        main_layout.addWidget(self.tab_widget)
        
        central_widget.setLayout(main_layout)
        self.setCentralWidget(central_widget)
        
        # Status Bar
        self.status_bar = QStatusBar()
        self.setStatusBar(self.status_bar)
        self.status_bar.showMessage(f"Bienvenido, {self.username.capitalize()} | {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
        
        # Timer para actualizar la hora en el status bar
        self.time_timer = QTimer(self)
        self.time_timer.timeout.connect(self.update_status_time)
        self.time_timer.start(1000)

    def update_status_time(self):
        """Actualiza la hora en la barra de estado."""
        self.status_bar.showMessage(f"Bienvenido, {self.username.capitalize()} | {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")

    def setup_tabs(self):
        """Configura los paneles de pestañas según los permisos del usuario."""
        
        # Pestaña 1: Dashboard (Visible para todos)
        dashboard_panel = DashboardPanel(self.business_logic)
        self.tab_widget.addTab(dashboard_panel, "🚀 Dashboard")
        
        # Pestaña 2: Gestión de Productos (Roles: admin, inventario, compras, finanzas, reportes)
        if self.user_role in ['admin', 'inventario', 'compras', 'finanzas', 'reportes']:
            product_panel = ProductManagementPanel(self.business_logic, self.user_role)
            self.tab_widget.addTab(product_panel, "📦 Productos/Inventario")
        
        # Pestaña 3: Gestión de Proveedores (Roles: admin, compras)
        if self.user_role in ['admin', 'compras', 'reportes']:
            proveedores_panel = ProveedoresManagementPanel(self.business_logic, self.user_role)
            self.tab_widget.addTab(proveedores_panel, "🚚 Proveedores")
        
        # Pestaña 4: Comprobantes (Roles: admin, compras, finanzas, reportes)
        if self.user_role in ['admin', 'compras', 'finanzas', 'reportes']:
            comprobantes_panel = ComprobantesPanel(self.business_logic, self.user_role)
            self.tab_widget.addTab(comprobantes_panel, "📑 Comprobantes")

    def show_welcome_message(self):
        """Muestra un mensaje de bienvenida personalizado al iniciar."""
        QMessageBox.information(
            self, 
            "Bienvenido al Sistema", 
            f"Hola, {self.username.capitalize()} ({self.user_role.upper()}).\n\n"
            f"Has iniciado sesión correctamente en el Sistema de Gestión de Alimentos Pro v{APP_VERSION}.\n\n"
            f"⚠️ **NOTA IMPORTANTE:** La conexión a la DB está configurada para usar la base de datos 'GestionAlimentos' con usuario 'root' y contraseña por defecto 'password123'.\n\n"
            f"Los usuarios de prueba disponibles son:\n"
            f"• paul, ana (admin)\n"
            f"• jhon (inventario), yessenia (compras), piero (finanzas), cassandra (reportes)\n"
            f"• vanina, miryam, natalia (atencion_cliente)\n\n"
            f"Todos usan la contraseña: admin123"
        )

# ==============================================================================
# CLASE DE APLICACIÓN PRINCIPAL (MAIN) - MEJORADA
# ==============================================================================

class LoginWindow(QDialog):
    """Ventana de login con diseño moderno."""
    
    def __init__(self, business_logic, parent=None):
        super().__init__(parent)
        self.business_logic = business_logic
        self.user_id = None
        self.username = None
        self.user_role = None
        self.error_message = None
        self.setup_ui()

    def setup_ui(self):
        self.setWindowTitle("🔐 Acceso al Sistema")
        self.setFixedSize(450, 600)
        
        # Paleta de colores para el login
        palette = QPalette()
        palette.setColor(QPalette.Window, QColor(COLORS['background']))
        self.setPalette(palette)
        
        main_layout = QVBoxLayout(self)
        main_layout.setContentsMargins(0, 0, 0, 0)
        
        # Header/Branding
        header_widget = QWidget()
        header_widget.setFixedHeight(150)
        header_widget.setStyleSheet(f"""
            background: qlineargradient(x1:0, y1:0, x2:1, y2:0, 
                                      stop:0 {COLORS['primary']}, 
                                      stop:1 {COLORS['secondary']});
            border-bottom-left-radius: 50px;
            border-bottom-right-radius: 50px;
        """)
        
        header_layout = QVBoxLayout(header_widget)
        header_layout.setAlignment(Qt.AlignCenter)
        
        logo_label = QLabel("🏢")
        logo_label.setStyleSheet("font-size: 36px; margin-bottom: -5px;")

        title_label = QLabel(APP_NAME)
        title_label.setStyleSheet("font-size: 18px; font-weight: 500; color: white; margin: 0;")

        subtitle_label = QLabel("Sistema Empresarial de Gestión")
        subtitle_label.setStyleSheet("font-size: 12px; color: rgba(255, 255, 255, 0.8); margin: 0;")

        header_layout.addWidget(logo_label)
        header_layout.addWidget(title_label)
        header_layout.addWidget(subtitle_label)
        main_layout.addWidget(header_widget)

        # Formulario de Login
        form_widget = QWidget()
        form_layout = QVBoxLayout(form_widget)
        form_layout.setContentsMargins(40, 30, 40, 30)
        form_layout.setSpacing(20)
        
        welcome_label = QLabel("Bienvenido de nuevo")
        welcome_label.setStyleSheet(f"font-size: 24px; font-weight: bold; color: {COLORS['text_primary']};")
        form_layout.addWidget(welcome_label, alignment=Qt.AlignCenter)
        
        self.username_input = ModernInput(placeholder="Nombre de Usuario")
        self.username_input.setMinimumHeight(50)
        self.username_input.setFocus()
        self.username_input.returnPressed.connect(self.login)
        form_layout.addWidget(self.username_input)
        
        self.password_input = ModernInput(placeholder="Contraseña")
        self.password_input.setEchoMode(QLineEdit.Password)
        self.password_input.setMinimumHeight(50)
        self.password_input.returnPressed.connect(self.login)
        form_layout.addWidget(self.password_input)
        
        self.login_btn = AnimatedButton("🔑 INICIAR SESIÓN")
        self.login_btn.setMinimumHeight(55)
        self.login_btn.clicked.connect(self.login)
        form_layout.addWidget(self.login_btn)
        
        # Enlace de ayuda/olvido
        help_label = QLabel(f"Versión: {APP_VERSION} | Soporte: contacto@empresa.com")
        help_label.setStyleSheet(f"font-size: 12px; color: {COLORS['text_secondary']};")
        form_layout.addWidget(help_label, alignment=Qt.AlignCenter)

        main_layout.addWidget(form_widget)
        main_layout.addStretch(1)

    def login(self):
        """Maneja la lógica de autenticación."""
        username = self.username_input.text().strip()
        password = self.password_input.text()

        if not username or not password:
            QMessageBox.warning(self, "Error de Login", "Por favor ingrese usuario y contraseña.")
            return

        self.login_btn.setEnabled(False)
        self.login_btn.setText("Verificando...")
        QApplication.processEvents() # Forzar la actualización de la UI

        try:
            self.user_id, self.username, self.user_role, self.error_message = self.business_logic.login_user(username, password)

            if self.user_id:
                self.accept()
            else:
                QMessageBox.critical(self, "Error de Login", self.error_message or "Usuario o contraseña incorrectos, o usuario inactivo.")
                self.password_input.clear()
                self.username_input.setFocus()

        except SecurityError as e:
            QMessageBox.critical(self, "Error de Seguridad", str(e))
        except DatabaseError as e:
            QMessageBox.critical(self, "Error de Base de Datos", f"No se pudo conectar al servidor. {str(e)}")
        except Exception as e:
            logging.error(f"Error inesperado en login: {traceback.format_exc()}")
            QMessageBox.critical(self, "Error Crítico", f"Ocurrió un error inesperado: {str(e)}")

        finally:
            self.login_btn.setEnabled(True)
            self.login_btn.setText("🔑 INICIAR SESIÓN")


class MainAppEmpresarial:
    """Clase principal que orquesta la aplicación."""
    
    def __init__(self):
        self.app = QApplication(sys.argv)
        self.config_manager = ConfigManager()
        self.db_manager = DatabaseManager(self.config_manager)
        self.security_manager = SecurityManager()
        self.business_logic = BusinessLogic(self.db_manager, self.security_manager)
        
        # Inicializar base de datos con usuarios por defecto si es necesario
        self._initialize_default_users()

    def _initialize_default_users(self):
        """Inicializa los usuarios por defecto si la tabla está vacía."""
        try:
            # 1. Verificar si la tabla de usuarios existe y si está vacía
            user_count_result = self.db_manager.execute_query(
                "SELECT COUNT(*) as count FROM usuarios",
                fetch_one=True
            )

            if user_count_result and user_count_result['count'] == 0:
                logging.warning("No se encontraron usuarios. Inicializando usuarios por defecto...")

                # 2. Generar hashes para los usuarios por defecto usando hash fijo para consistencia
                fixed_hash = "$2b$12$LQv3c1yqBzwSxJ4T4YwZeuY6n6b9s8JZ8VkQdE7fM5rNkYbLd8HKO"  # Hash de 'admin123'

                users_to_insert = []
                for username, password, role in USUARIOS_SISTEMA:
                    users_to_insert.append((username, fixed_hash, role))

                # 3. Inserción masiva
                conn = self.db_manager.connect()
                with conn.cursor() as cur:
                    insert_query = """
                    INSERT INTO usuarios
                    (nombreusuario, contrasena, rol, activo, fecha_creacion)
                    VALUES (%s, %s, %s, TRUE, NOW())
                    """
                    cur.executemany(insert_query, users_to_insert)
                    conn.commit()

                logging.info(f"Se insertaron {len(users_to_insert)} usuarios por defecto correctamente.")

        except DatabaseError as e:
            logging.error(f"Error en la BD durante la inicialización de usuarios: {e}. La tabla 'usuarios' podría no existir.")
        except Exception as e:
            logging.critical(f"Error crítico durante la inicialización de usuarios: {traceback.format_exc()}")

    def run(self):
        """Ejecuta la secuencia de inicio de la aplicación."""
        try:
            # 1. Iniciar ventana de Login
            login_window = LoginWindow(self.business_logic)
            login_result = login_window.exec_()
            
            if login_result == QDialog.Accepted:
                # 2. Iniciar ventana principal
                main_window = ModernMainWindow(
                    self.business_logic,
                    login_window.user_id,
                    login_window.username,
                    login_window.user_role,
                    self.config_manager
                )
                
                # Centrar ventana
                screen = QDesktopWidget().screenGeometry()
                main_window.move(
                    (screen.width() - main_window.width()) // 2,
                    (screen.height() - main_window.height()) // 2
                )
                
                main_window.show()
                
                logging.info("Aplicación iniciada correctamente")
                return self.app.exec_()
            else:
                logging.info("Aplicación cerrada por el usuario")
                return 0
                
        except Exception as e:
            logging.critical(f"Error crítico al iniciar la aplicación: {traceback.format_exc()}")
            QMessageBox.critical(
                None,
                "Error Crítico - No se puede iniciar la aplicación",
                f"Error: {str(e)}\n\n"
                "🔍 **Diagnóstico Profesional:**\n\n"
                "Este error indica problemas de conectividad con la base de datos MySQL.\n\n"
                "📋 **Pasos de Solución Recomendados:**\n\n"
                "1. **Verificar MySQL Server:**\n"
                "   • Abra Terminal y ejecute: brew services list\n"
                "   • Busque 'mysql' y verifique que esté 'started'\n"
                "   • Si no está corriendo: brew services start mysql\n\n"
                "2. **Verificar Base de Datos:**\n"
                "   • Conéctese a MySQL: mysql -u root -p\n"
                "   • Ejecute: SHOW DATABASES;\n"
                "   • Busque 'GestionAlimentos' en la lista\n"
                "   • Si no existe: CREATE DATABASE GestionAlimentos;\n\n"
                "3. **Verificar Credenciales:**\n"
                "   • Usuario: root\n"
                "   • Contraseña: (vacía por defecto o 'password123')\n"
                "   • Host: localhost\n"
                "   • Puerto: 3306\n\n"
                "4. **Importar Estructura de BD:**\n"
                "   • Use el archivo GestionAlimentos.sql incluido\n"
                "   • Ejecute: mysql -u root GestionAlimentos < GestionAlimentos.sql\n\n"
                "5. **Probar Conexión:**\n"
                "   • Ejecute: python test_connection.py\n\n"
                f"📄 **Detalles Técnicos:** {os.path.abspath('app_errors.log')}\n\n"
                "💡 **Soporte:** Si persiste el problema, contacte al administrador del sistema."
            )
            return 1
        finally:
            try:
                self.db_manager.close()
            except:
                pass

# ==============================================================================
# EJECUCIÓN PRINCIPAL
# ==============================================================================

if __name__ == "__main__":
    # Establecer estilo empresarial
    QApplication.setStyle("Fusion")
    
    # Crear y ejecutar aplicación
    app = MainAppEmpresarial()
    
    # Ejecutar aplicación
    exit_code = app.run()
    
    # Salir con el código apropiado
    sys.exit(exit_code)